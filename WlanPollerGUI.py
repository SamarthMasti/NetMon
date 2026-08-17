# ==========================================================
# WlanPollerGUI.py  (Updated)
# ==========================================================
# Author: generated/fixed by assistant
# Purpose: Complete GUI for CISCO WLAN POLLER/PARSER (PySide6)
# ==========================================================

import os
import sys
import socket
import re
from datetime import datetime
import time
from typing import List, Optional
from pathlib import Path

from PySide6.QtWidgets import QHeaderView
from PySide6.QtCore import Qt, QSize, QThread, QTimer, Signal
from PySide6.QtGui import QFont, QPixmap, QIcon
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton, QTextEdit, QLineEdit,
    QVBoxLayout, QHBoxLayout, QGridLayout, QComboBox, QCheckBox, QListWidget,
    QStackedWidget, QMessageBox, QGroupBox, QTableWidget, QTableWidgetItem,
    QFileDialog, QSizePolicy, QSpacerItem, QFrame, QProgressBar, QFormLayout,
    QRadioButton, QButtonGroup
)
from PollerEngine import PollerEngine
from PollerEngine import decrypt_value, encrypt_value
from PySide6.QtGui import QColor
from PollerEngine import DATAPATH_MON_DEFAULT_ITERATIONS, DATAPATH_MON_DEFAULT_INTERVAL_SEC
APP_NAME = "CISCO WLAN POLLER GUI"
APP_VERSION = "v5.10"
try:
    from ApFlashVulnerableChecker import analyze_logs
except ImportError as e:
    print("ApFlashVulnerableChecker import failed:", e)
    analyze_logs = None

try:
    from PollerEngine import PollerEngine
except ImportError as e:
    raise ImportError(f"CRITICAL: Failed to import PollerEngine module: {e}")

def get_app_base_dir() -> Path:
    """
    Returns directory where:
    - WlanPollerGUI.app lives (macOS)
    - WlanPollerGUI.exe lives (Windows)
    - script folder when running source
    """

    if getattr(sys, "frozen", False):

        exe = Path(sys.executable).resolve()

        # macOS bundled app
        if exe.parent.name == "MacOS" and exe.parent.parent.name == "Contents":
            return exe.parents[3]   # <-- outside .app

        return exe.parent

    return Path(__file__).resolve().parent

BASE_DIR = get_app_base_dir()

DATA_DIR = BASE_DIR / "data"
CONFD_DIR = BASE_DIR / "confd"

DATA_DIR.mkdir(parents=True, exist_ok=True)
CONFD_DIR.mkdir(parents=True, exist_ok=True)



CONFIG_FILE = str(CONFD_DIR / "config.ini")

CONFD = str(CONFD_DIR)
# Optional Excel export
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font as XLFont, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None

# ---------------- Visual constants ----------------
SIDEBAR_BG = "#000000"
HERO_START = "#000000"
HERO_END = "#000000"
CARD_BG = "#ffffff"
CARD_BORDER = "#e6e8eb"
ACCENT = "#16a34a"
TEXT_PRIMARY = "#0f1724"
TEXT_MUTED = "#6b7280"

FONT_BODY = QFont("Roboto", 11)
FONT_TITLE = QFont("Roboto", 22, QFont.Weight.Bold)
FONT_CARD_TITLE = QFont("Roboto", 18, QFont.Weight.DemiBold)


def apply_global_style(app: QApplication):
    """
    Apply global QSS. Uses an f-string triple-quoted string so the CSS
    is a proper Python string literal and can reference color constants.
    """
    app.setFont(FONT_BODY)

    qss = f"""
    QWidget {{
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #f8fafc, stop:1 #eef2f7);
        color: {TEXT_PRIMARY};
        font-family: Roboto, "Open Sans", "Segoe UI", Arial, sans-serif;
    }}

    QLabel#heroTitle {{
        color: white;
    }}

    QListWidget {{
        background: {SIDEBAR_BG};
        color: white;
        border: none;
        padding-top: 10px;
    }}
    QListWidget::item {{
        padding: 12px 18px;
        border-radius: 6px;
    }}
    QListWidget::item:selected {{
        background: #1c1c1c;
    }}

    QGroupBox {{
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 10px;

        margin-top: 20px;
        padding-top: 18px;  

    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        subcontrol-position: top left;
        left: 16px;
        top: 6px;
        padding: 0px 6px;
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 10px;
    }}

    QPushButton {{
        background: #000000;
        color: white;
        border-radius: 8px;
        padding: 8px 14px;
        min-width: 110px;
    }}
    QPushButton:hover {{
            background: #1c1c1c;
    }}

    QPushButton:pressed {{
        background: #111111;
    }}

    QPushButton:disabled {{
        background: #d1d5db;
        color: #9ca3af;
    }}


    /* ---------------------------
       NAV buttons: black like sidebar
       Mark a button with: btn.setProperty("nav", True)
       --------------------------- */

    QTextEdit, QLineEdit, QComboBox {{
        background: #ffffff;
        border: 1px solid #e6e8eb;
        border-radius: 6px;
        padding: 6px;
    }}

    QTableWidget {{
    background: white;
    border: 1px solid #d1d5db;          /* outer border */
    border-radius: 10px;
    gridline-color: #e5e7eb;
    selection-background-color: #dcfce7;
    selection-color: #000000;
    }}

    QTableWidget::item {{
    border-right: 1px solid #f1f5f9;    /* subtle vertical lines */
    border-bottom: 1px solid #f1f5f9;   /* subtle row lines */
    padding: 10px;
    }}

    QHeaderView::section {{
    background-color: #f3f4f6;
    border-right: 1px solid #e5e7eb;
    border-bottom: 1px solid #d1d5db;
    padding: 8px;
    font-weight: 600;
     }}
    QProgressBar {{
        text-align: center;
        font-weight: 700;
        color: #000000; /* ensure percent text is black */
    }}
    """

    app.setStyleSheet(qss)


def safe_pixmap(path: str, size: Optional[QSize] = None) -> Optional[QPixmap]:
    if not path or not os.path.exists(path):
        return None
    pix = QPixmap(path)
    if size:
        return pix.scaled(size, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
    return pix


def is_ipv4_or_ipv6(addr: str) -> bool:
    addr = addr.strip()
    if not addr:
        return False
    try:
        socket.inet_pton(socket.AF_INET, addr)
        return True
    except Exception:
        pass
    try:
        socket.inet_pton(socket.AF_INET6, addr)
        return True
    except Exception:
        return False
def normalize_ap_entry(parts):
    """
    Normalize AP entry into (ip, model, name)
    Handles all cases safely.
    """

    ip = parts[0] if len(parts) >= 1 else ""
    model = ""
    name = ""

    if len(parts) >= 3:
        model = parts[1]
        name = " ".join(parts[2:]).strip()

    elif len(parts) == 2:
        val = parts[1]

        # Detect if it's a Cisco model (starts with C, AIR, CW etc.)
        if re.match(r"^(C\d+|AIR-|CW)", val, re.IGNORECASE):
            model = val
            name = ""
        else:
            name = val
            model = ""

    # ---- Extract model from name if missing ----
    if not model and name:
        match = re.search(r"(C\d+\w*|AIR-[A-Z0-9-]+|CW\d+\w*)", name, re.IGNORECASE)
        if match:
            model = match.group(1)

    # ---- Final fallbacks ----
    if not name and ip:
        name = f"AP_{ip.replace('.', '_')}"

    if not model:
        model = "UNKNOWN"

    return ip, model, name

import configparser
from dataclasses import dataclass



# Writable location for macOS bundled app
from pathlib import Path
import sys
import os


from PollerEngine import IniStore
'''
def __init__(self, path: str):
        self.path = path
        self.cfg = configparser.ConfigParser(interpolation=None)
        if os.path.exists(path):
            self.cfg.read(path)
        if self.cfg.has_option("WLC", "wlcipaddr"):
            val = self.cfg.get("WLC", "wlcipaddr")
            self.cfg.set("WLC", "wlc_ip", val)
            self.cfg.remove_option("WLC", "wlcipaddr")

def get(self, section: str, key: str, default: str = "") -> str:
        val = self.cfg.get(section, key, fallback=default)

        if "pasw" in key.lower() or "password" in key.lower() or "enable" in key.lower():
            return decrypt_value(val)

        return val

def bulk_set(self, section: str, data: dict):
        if not self.cfg.has_section(section):
            self.cfg.add_section(section)
        for k, v in data.items():
            if "pasw" in k.lower() or "password" in k.lower() or "enable" in k.lower():
                if isinstance(v, str) and not v.startswith("ENC::"):
                    v = encrypt_value(v)
            self.cfg.set(section, k, v)

def save(self):
        Path(self.path).parent.mkdir(parents=True, exist_ok=True)
        with open(self.path, "w", encoding="utf-8") as f:
            self.cfg.write(f)

'''
@dataclass
class ApRow:
    ip: str
    model: str
    name: str
    site_tag: str = ""
    wlc_ip: str = ""


# ---------------- Worker ----------------

class PollerWorker(QThread):
    progress = Signal(int)
    log = Signal(str)
    ap_update = Signal(int, str, str, str, str,str)
    finished_ok = Signal(dict)
    failed = Signal(str)
    # Bulk WLC mode only (>3 WLCs): reports (completed, total) WLC count as
    # each WLC finishes. Never emitted for the manual (<=3 WLC) path.
    wlc_progress = Signal(int, int)

    def _engine_progress(self, pct):
        self.progress_sig.emit(pct)

    def _engine_log(self, msg):
        self.log_sig.emit(str(msg))



    def __init__(
            self,
            operation_type: str,
            workflow: str,
            wlc_cmds: List[str],
            ap_cmds: List[str],
            ap_filter_mode: str,
            site_tag: str,
            model_group: str,
            ap_device: str,
            ap_list_file: str = "",
            ap_mode: str = "AP Custom Cmd List",
            iterations_enabled: bool = False,
            iteration_count:    int  = 1,
            iteration_interval: int  = 300,
            client_mac: str = "", 
    ):
        super().__init__()
        self.operation_type = operation_type
        self.workflow = workflow
        self.wlc_cmds = wlc_cmds
        self.ap_cmds = ap_cmds
        self.ap_filter_mode = ap_filter_mode
        self.site_tag = site_tag
        self.model_group = model_group
        self.ap_device = ap_device
        self.ap_list_file = ap_list_file
        self.ap_mode = ap_mode
        self.iterations_enabled  = iterations_enabled
        self.iteration_count     = max(1, min(50, iteration_count))
        self.iteration_interval  = max(0, min(18000, iteration_interval))
        self.client_mac = client_mac           # NEW
    def run(self):
        # AP Datapath Queue Mon manages its own internal iteration loop
        # (single SSH session, reused across iterations) — never let the
        # outer per-run reconnect loop wrap it.
        if self.workflow == "AP Datapath Queue Mon":
            iterations = 1
        else:
            iterations = self.iteration_count if self.iterations_enabled else 1

        for _iter in range(iterations):

            # ── Iteration header ─────────────────────────────────
            if iterations > 1:
                self.log.emit(f"\n{'='*54}")
                self.log.emit(f"[ITERATION] Run {_iter + 1} of {iterations}")
                self.log.emit(f"{'='*54}\n")

            # ── Existing run body (unchanged, indented one level) ─
            engine = None
        
            try:
                # create engine inside try/except so creation failures are visible
                try:

                    engine = PollerEngine(
                        log_cb=lambda msg: self.log.emit(msg),
                        progress_cb=lambda pct: self.progress.emit(pct),
                    ap_update_cb=lambda i, ip, model, status, name, wlc:
                        self.ap_update.emit(i, ip, model, status, name, wlc)
                    )
                    engine.operation = self.operation_type

                    # Only pass workflow to engine when WLC is involved
                    if self.operation_type == "WLC & AP":
                        engine.workflow = self.workflow
                    else:
                        engine.workflow = ""

                except Exception as e:
                    try:
                        self.log.emit(f"PollerWorker: engine creation failed: {e}")
                    except Exception:
                        pass
                    try:
                        self.failed.emit(str(e))
                    except Exception:
                        pass
                    continue

                try:
                    self.log.emit("PollerWorker: engine created, starting operation")
                except Exception:
                    pass

                start = datetime.now()
                summary = {"start": start, "operation": self.operation_type}

                # --- WLC Only ---
                if self.operation_type == "WLC Only":

                    # 🔥 NEW WORKFLOW
                    if self.workflow == "Client Stuck In Auth Loop":

                        engine.enable_debug_collection = getattr(self, "enable_debug_collection", False)
                        delete_list = engine.run_client_auth_workflow()
                        
                        

                        summary["delete_list"] = delete_list
                        summary["clients_detected"] = len(delete_list)
                        summary["data_dir"] = engine.data_dir   # ← ADD THIS LINE
                        self.log.emit("")
                        self.log.emit("=" * 50)
                        self.log.emit(" CLIENT AUTH LOOP SUMMARY ")
                        self.log.emit("=" * 50)
                        self.log.emit(f" Total stuck clients detected: {len(delete_list)}")

                        if delete_list:
                            for mac in delete_list:
                                self.log.emit(f" Deauthenticated: {mac}")
                        else:
                            self.log.emit(" No clients stuck in auth loop")

                        self.log.emit("=" * 50)

                        summary["end"] = datetime.now()
                        self.finished_ok.emit(summary)
                        return

                    # OLD FLOW
                    
                    if not self.wlc_cmds:
                        raise ValueError("WLC Cmd List is empty.")

                    out = engine.run_wlc_cmds(self.wlc_cmds)
                    summary.update({"wlc_output": out})
                    summary["end"] = datetime.now()
                    self.finished_ok.emit(summary)
                    

                # --- AP Only ---
                elif self.operation_type == "AP Only":

                    if not self.ap_list_file or not os.path.exists(self.ap_list_file):
                        raise ValueError(f"AP list file missing: {self.ap_list_file}")

                    ap_rows = []

                    with open(self.ap_list_file, "r", encoding="utf-8", errors="ignore") as f:
                        for line in f:
                            s = line.strip()
                            if not s:
                                continue

                            parts = [p.strip() for p in (s.split(",") if "," in s else s.split())]

                            ip, model, name = normalize_ap_entry(parts)

                            if ip:
                                ap_rows.append(ApRow(ip=ip, model=model, name=name))
                    if not ap_rows:
                        raise ValueError("AP list file is empty.")

                    if self.workflow == "AP Datapath Queue Mon":
                        results = engine.run_ap_datapath_queue_monitor(ap_rows)
                        summary.update({
                            "ap_total": len(ap_rows),
                            "ap_success": engine.success,
                            "ap_failed": engine.failed,
                            "data_dir": getattr(engine, "data_dir", ""),
                            "workflow": self.workflow,
                            "datapath_results": results,
                        })
                        self.log.emit("")
                        self.log.emit("=" * 56)
                        self.log.emit("  AP DATAPATH QUEUE MONITOR SUMMARY")
                        self.log.emit("=" * 56)
                        for r in results:
                            self.log.emit(f"  {r['ap_name']} ({r['ap_ip']}): {r.get('status')}")
                            for dp in r.get("datapaths", []):
                                self.log.emit(
                                    f"      [{dp.get('datapath_id')}] clients={dp.get('clients')} -> "
                                    f"{dp.get('overall_assessment', '')}"
                                )
                                if dp.get("recommended_recovery"):
                                    self.log.emit(f"          Recovery: {dp['recommended_recovery']}")
                        self.log.emit("=" * 56)
                        self.log.emit(f"  Total APs   : {len(ap_rows)}")
                        self.log.emit(f"  Success     : {engine.success}")
                        self.log.emit(f"  Failed      : {engine.failed}")
                        self.log.emit("=" * 56)
                        # ---- NEW: Status Run Check file when APs failed enable mode ----
                        enable_failed = getattr(engine, "enable_failed_aps", [])
                        if enable_failed and analyze_logs:
                            vuln_rows, _ = analyze_logs(
                                str(summary["data_dir"]),
                                enable_failed_aps=enable_failed
                            )
                            summary["vulnerable_rows"] = vuln_rows
                            self.log.emit(
                                f"[DATAPATH] {len(vuln_rows)} AP(s) flagged for enable-mode failure status check."
                            )

                        summary["end"] = datetime.now()
                        self.finished_ok.emit(summary)
                        return

                    # ---- All other AP Only workflows (unchanged) ----
                    if not self.ap_cmds:
                        raise ValueError("AP Cmd List is empty.")

                    self.log.emit(f"[DEBUG] Parsed AP rows: {len(ap_rows)}")
                    self.log.emit(f"[DEBUG] AP list file path = {self.ap_list_file}")
                    engine.run_ap_poller(ap_rows, self.ap_device, self.ap_cmds, ap_mode=self.ap_mode)

                    summary.update({
                        "ap_total": len(ap_rows),
                        "ap_success": engine.success,
                        "ap_failed": engine.failed,
                        "data_dir": getattr(engine, "data_dir", ""),
                        "workflow": self.workflow
                    })

                    # AP COUNT SUMMARY
                    self.log.emit("")
                    self.log.emit("=" * 56)
                    self.log.emit("  AP COUNT SUMMARY")
                    self.log.emit("=" * 56)
                    self.log.emit(f"  Total APs in file   : {len(ap_rows)}")
                    self.log.emit(f"  APs processed       : {len(ap_rows)}")
                    self.log.emit(f"  Success             : {engine.success}")
                    self.log.emit(f"  Failed              : {engine.failed}")
                    self.log.emit("=" * 56)

                    # THEN vulnerability analysis
                    if (
                        analyze_logs
                        and not (getattr(self, "enable_tmp_cleanup", False) or getattr(self, "enable_reload", False))
                        and (self.workflow == "AP Flash Checker" or getattr(engine, "enable_failed_aps", []))
                    ):
                        self.log.emit("")
                        self.log.emit("=" * 56)
                        if getattr(self, "test_after_iteration", False):
                            try:
                                deletetest_dir = os.path.join(BASE_DIR, "deletetest")
                                if os.path.isdir(deletetest_dir):
                                    copied = 0
                                    for fn in os.listdir(deletetest_dir):
                                        src = os.path.join(deletetest_dir, fn)
                                        if os.path.isfile(src):
                                            import shutil
                                            shutil.copy2(src, os.path.join(str(summary["data_dir"]), fn))
                                            copied += 1
                                    self.log.emit(f"[TEST MODE] Copied {copied} file(s) from deletetest/ into run folder for parsing.")
                                else:
                                    self.log.emit("[TEST MODE] deletetest/ folder not found — skipping.")
                            except Exception as _e:
                                self.log.emit(f"[TEST MODE] deletetest copy failed: {_e}")
                        self.log.emit("  RUNNING FLASH SUSCEPTIBILITY ANALYSIS...")
                        self.log.emit("  Please wait — scanning AP output logs.")
                        self.log.emit("=" * 56)
                        self.progress.emit(0)
                        vuln_rows, _ = analyze_logs(
                            str(summary["data_dir"]),
                            enable_failed_aps=getattr(engine, "enable_failed_aps", [])
                        )
                        self.log.emit(f"[FLASH DEBUG] Parser input folder = {summary['data_dir']}")

                        try:
                            for f in os.listdir(summary["data_dir"]):
                                self.log.emit(f"[FLASH DEBUG] Found file = {f}")
                        except Exception as e:
                            self.log.emit(f"[FLASH DEBUG] Directory read failed: {e}")

                        summary["vulnerable_rows"] = vuln_rows
                        self.log.emit(f"  Susceptibility scan complete. Found: {len(vuln_rows)} Susceptible AP(s)")
                        self.log.emit("=" * 56)
                        self.progress.emit(100)

                    # ---- NEW: fold in enable-mode-failed APs regardless of workflow ----
                    enable_failed = getattr(engine, "enable_failed_aps", [])
                    if enable_failed:
                        vuln_rows_existing = summary.get("vulnerable_rows", [])
                        existing_ips = {vr.get("ap_ip") for vr in vuln_rows_existing}
                        ENABLE_FAIL_MSG = (
                            "Failed to enter enable mode. This may be caused by high "
                            "storage utilization on the active boot partition or a "
                            "missing 'secret' parameter. Please ensure the 'secret' "
                            "parameter is provided and retry."
                        )
                        added = 0
                        for ap_fail in enable_failed:
                            if ap_fail.get("ip") in existing_ips:
                                continue
                            vuln_rows_existing.append({
                                "ap_name": ap_fail.get("name", ""),
                                "ap_model": ap_fail.get("model", "") or "UNKNOWN",
                                "ap_ip": ap_fail.get("ip", ""),
                                "recovery": ENABLE_FAIL_MSG,
                                "active_boot_part": "Unknown",
                                "partition_note": "Enable mode failed — flash status could not be verified.",
                            })
                            added += 1
                        summary["vulnerable_rows"] = vuln_rows_existing
                        if added:
                            self.log.emit(
                                f"[FLASH DEBUG] Added {added} AP(s) with 'Failed to enter "
                                f"enable mode' to Susceptible Table."
                            )

                    summary["end"] = datetime.now()
                    self.finished_ok.emit(summary)
                    
                # --- WLC & AP ---
                elif self.operation_type == "WLC & AP":
                    import threading
                    from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed

                    # Safety redirect: if cronjob saved wrong operation_type, catch it here
                    if self.workflow == "Client Stuck In Auth Loop":
                        engine.enable_debug_collection = getattr(self, "enable_debug_collection", False)
                        delete_list = engine.run_client_auth_workflow()
                        summary["delete_list"] = delete_list
                        summary["clients_detected"] = len(delete_list)
                        summary["data_dir"] = engine.data_dir   # ← ADD THIS LINE
                        self.log.emit("")
                        self.log.emit("=" * 50)
                        self.log.emit(" CLIENT AUTH LOOP SUMMARY ")
                        self.log.emit("=" * 50)
                        self.log.emit(f" Total stuck clients detected: {len(delete_list)}")
                        if delete_list:
                            for mac in delete_list:
                                self.log.emit(f" Deauthenticated: {mac}")
                        else:
                            self.log.emit(" No clients stuck in auth loop")
                        self.log.emit("=" * 50)
                        summary["end"] = datetime.now()
                        self.finished_ok.emit(summary)
                        return

                    # Bulk WLC mode (>3 WLCs, uploaded via Excel/text): use the
                    # uploaded list + shared credentials instead of config.ini
                    # sections. Manual (<=3 WLC) runs are completely unaffected —
                    # wlc_bulk_list stays empty for them and this branch is skipped.
                    wlc_bulk_list = getattr(self, "wlc_bulk_list", None) or []
                    if wlc_bulk_list:
                        wlc_sections = engine.register_bulk_wlc_sections(
                            wlc_bulk_list,
                            getattr(self, "bulk_wlc_user", ""),
                            getattr(self, "bulk_wlc_pasw", ""),
                        )
                    else:
                        wlc_sections = engine._get_wlc_sections_list()
                    # ── Headless guard: verify ap_cmds are not empty before starting ──
                    if not self.ap_cmds and self.workflow != "AP Datapath Queue Mon":
                        raise ValueError(
                            "[WLC & AP] AP Cmd List is empty. "
                            "Check that run_profile.json contains ap_cmds and the profile was saved correctly."
                        )
                    self.log.emit(f"[WORKER] WLC sections found: {wlc_sections}")
                    self.log.emit(f"[WORKER] AP commands to run: {self.ap_cmds}")
                    self.log.emit(f"[WORKER] WLC commands to run: {self.wlc_cmds}")
                    if not wlc_sections:
                        raise ValueError(
                            "[WLC & AP] No WLC sections found in config.ini. "
                            "Credentials may not have been saved correctly."
                        )
                    all_filtered = []
                    all_filtered_lock = threading.Lock()
                    _summary_lock = threading.Lock()

                    total_success = 0
                    total_failed = 0
                    _count_lock = threading.Lock()

                    def _process_one_wlc(section):
                        """Fetch APs from one WLC, apply filters, then immediately poll those APs."""

                        # ---- Step 1: Run WLC commands ----
                        if self.wlc_cmds:
                            engine._run_wlc_cmds_for_section(section, self.wlc_cmds)
                    
                        # ---- Step 2: Fetch AP list ----
                        rows = engine._fetch_ap_list_for_section(section)

                        # ---- Step 3: Apply filters ----
                        local_total = len(rows)
                        site_tag_used = ""

                        if self.ap_filter_mode == "SITE":
                            rows, local_total = engine._filter_by_site_tag_section(
                                section, rows, self.site_tag
                            )
                            site_tag_used = self.site_tag
                        elif self.ap_filter_mode == "MODEL":
                            rows = engine.filter_by_model_group(rows, self.model_group)

                        with _summary_lock:
                            if self.ap_filter_mode == "SITE":
                                summary["TotalApCnt"] = summary.get("TotalApCnt", 0) + local_total
                                summary["SiteTagNameFilter"] = self.site_tag

                        if not rows:
                            self.log.emit(
                                f"[WORKER] {section}: AP list is EMPTY. "
                                f"'show ap summary' returned 0 APs or WLC SSH failed. "
                                f"Check WLC connectivity and credentials in config.ini."
                            )
                            return 0, 0

                        # ---- Step 4: Write filtered list (thread-safe append) ----
                        with all_filtered_lock:
                            all_filtered.extend(rows)

                        # ---- Step 5: Poll APs for THIS WLC immediately (parallel within WLC) ----
                        if self.workflow == "AP Datapath Queue Mon":          # ← ADD
                            return 0, 0
                        self.log.emit(
                            f"[WORKER] {section}: starting AP polling for {len(rows)} APs..."
                        )
                        s, f = engine.run_ap_poller(rows, self.ap_device, self.ap_cmds)
                        self.log.emit(
                            f"[WORKER] {section}: AP polling done. Success={s} Failed={f}"
                        )
                        return s, f

                    # ---- Run all WLCs in parallel ----
                    self.log.emit(
                        f"[WORKER] Starting {len(wlc_sections)} WLC(s) in parallel..."
                    )

                    # Cap concurrent WLC connections at 10. For manual mode
                    # (<=3 WLCs) this is a no-op — min(3, 10) == 3, identical
                    # to the previous behavior. Only bulk mode (up to 50 WLCs)
                    # is actually throttled by this.
                    wlc_max_workers = min(len(wlc_sections), 10)
                    wlc_done_count = 0

                    with ThreadPoolExecutor(max_workers=wlc_max_workers) as wlc_executor:
                        wlc_futures = {
                            wlc_executor.submit(_process_one_wlc, sec): sec
                            for sec in wlc_sections
                        }
                        for fut in _as_completed(wlc_futures):
                            sec = wlc_futures[fut]
                            try:
                                s, f = fut.result()
                                with _count_lock:
                                    total_success += s
                                    total_failed += f
                            except Exception as e:
                                self.log.emit(f"[WORKER] {sec} failed: {e}")
                            finally:
                                if wlc_bulk_list:
                                    with _count_lock:
                                        wlc_done_count += 1
                                        done_snapshot = wlc_done_count
                                    self.wlc_progress.emit(done_snapshot, len(wlc_sections))

                    # ---- Write combined filtered list after all WLCs done ----
                    if all_filtered:
                        engine.write_filtered_ap_list(all_filtered)

                    if not all_filtered:
                        self.log.emit("[WORKER] Warning: Filtered AP list is empty — no APs to poll.")
                        summary.update({"ap_total": 0, "ap_success": 0, "ap_failed": 0,
                                        "data_dir": engine.data_dir})
                        summary["end"] = datetime.now()
                        self.finished_ok.emit(summary)
                        
                    elif self.workflow == "AP Datapath Queue Mon":            # ← ADD THIS WHOLE BLOCK
                        self.log.emit(f"[WORKER] Starting AP Datapath Queue Monitor across {len(all_filtered)} AP(s)...")
                        results = engine.run_ap_datapath_queue_monitor(all_filtered)
                        summary.update({
                            "datapath_results": results,
                            "workflow": self.workflow,
                        })
                        self.log.emit("")
                        self.log.emit("=" * 56)
                        self.log.emit("  AP DATAPATH QUEUE MONITOR SUMMARY")
                        self.log.emit("=" * 56)
                        for r in results:
                            self.log.emit(f"  {r['ap_name']} ({r['ap_ip']}): {r.get('status')}")
                            for dp in r.get("datapaths", []):
                                self.log.emit(
                                    f"      [{dp.get('datapath_id')}] clients={dp.get('clients')} -> "
                                    f"{dp.get('overall_assessment', '')}"
                                )
                                if dp.get("recommended_recovery"):
                                    self.log.emit(f"          Recovery: {dp['recommended_recovery']}")
                        self.log.emit("=" * 56)
                        total_success = engine.success
                        total_failed = engine.failed
                        # ---- NEW: Status Run Check file when APs failed enable mode ----
                        enable_failed = getattr(engine, "enable_failed_aps", [])
                        if enable_failed and analyze_logs:
                            vuln_rows, _ = analyze_logs(
                                str(engine.data_dir),
                                enable_failed_aps=enable_failed
                            )
                            summary["vulnerable_rows"] = vuln_rows
                            self.log.emit(
                                f"[DATAPATH] {len(vuln_rows)} AP(s) flagged for enable-mode failure status check."
                            )
                    # ---- Flash checker analysis (also runs for any workflow when enable-mode failures occurred) ----
                    elif (analyze_logs and
                            not (getattr(self, "enable_tmp_cleanup", False) or
                                getattr(self, "enable_reload", False)) and
                            (self.workflow == "AP Flash Checker" or getattr(engine, "enable_failed_aps", []))):
                        self.log.emit("=" * 56)
                        if getattr(self, "test_after_iteration", False):
                            try:
                                deletetest_dir = os.path.join(BASE_DIR, "deletetest")
                                if os.path.isdir(deletetest_dir):
                                    copied = 0
                                    for fn in os.listdir(deletetest_dir):
                                        src = os.path.join(deletetest_dir, fn)
                                        if os.path.isfile(src):
                                            import shutil
                                            shutil.copy2(src, os.path.join(str(engine.data_dir), fn))
                                            copied += 1
                                    self.log.emit(f"[TEST MODE] Copied {copied} file(s) from deletetest/ into run folder for parsing.")
                                else:
                                    self.log.emit("[TEST MODE] deletetest/ folder not found — skipping.")
                            except Exception as _e:
                                self.log.emit(f"[TEST MODE] deletetest copy failed: {_e}")
                        self.log.emit(" RUNNING FLASH SUSCEPTIBILITY ANALYSIS...")
                        self.progress.emit(0)
                        vuln_rows, _ = analyze_logs(
                            str(engine.data_dir),
                            enable_failed_aps=getattr(engine, "enable_failed_aps", [])
                        )

                        summary["vulnerable_rows"] = vuln_rows
                        self.log.emit(
                            f"  Scan complete. Found: {len(vuln_rows)} susceptible AP(s)"
                        )
                        self.progress.emit(100)

                    # ---- NEW: fold in enable-mode-failed APs regardless of workflow ----
                    enable_failed = getattr(engine, "enable_failed_aps", [])
                    if enable_failed:
                        vuln_rows_existing = summary.get("vulnerable_rows", [])
                        existing_ips = {vr.get("ap_ip") for vr in vuln_rows_existing}
                        ENABLE_FAIL_MSG = (
                            "Failed to enter enable mode. This may be caused by high "
                            "storage utilization on the active boot partition or a "
                            "missing 'secret' parameter. Please ensure the 'secret' "
                            "parameter is provided and retry."
                        )
                        added = 0
                        for ap_fail in enable_failed:
                            if ap_fail.get("ip") in existing_ips:
                                continue
                            vuln_rows_existing.append({
                                "ap_name": ap_fail.get("name", ""),
                                "ap_model": ap_fail.get("model", "") or "UNKNOWN",
                                "ap_ip": ap_fail.get("ip", ""),
                                "recovery": ENABLE_FAIL_MSG,
                                "active_boot_part": "Unknown",
                                "partition_note": "Enable mode failed — flash status could not be verified.",
                            })
                            added += 1
                        summary["vulnerable_rows"] = vuln_rows_existing
                        if added:
                            self.log.emit(
                                f"[FLASH DEBUG] Added {added} AP(s) with 'Failed to enter "
                                f"enable mode' to Susceptible Table."
                            )

                    summary.update({
                        "ap_total": len(all_filtered),
                        "ap_success": total_success,
                        "ap_failed": total_failed,
                        "data_dir": engine.data_dir,
                        "TotalApCnt": summary.get("TotalApCnt", len(all_filtered)),
                    })
                    summary["end"] = datetime.now()
                    self.finished_ok.emit(summary)

                    
                else:
                    raise ValueError("Unknown operation.")


            except Exception as e:

                import traceback

                traceback.print_exc()

                try:

                    self.log.emit(f"[ERROR] {str(e)}")

                except Exception:

                    pass

                try:

                    self.failed.emit(str(e))

                except Exception:

                    pass



            finally:
                # cleanup engine if it exposes shutdown/close
                try:
                    if engine is not None:
                        if hasattr(engine, "shutdown") and callable(getattr(engine, "shutdown")):
                            try:
                                engine.shutdown()
                            except Exception:
                                pass
                        if hasattr(engine, "close") and callable(getattr(engine, "close")):
                            try:
                                engine.close()
                            except Exception:
                                pass
                except Exception:
                    pass
            # ── Inter-iteration sleep (skip after last run) ───────
            if _iter < iterations - 1:
                wait = self.iteration_interval
                self.log.emit(
                    f"\n[ITERATION] Run {_iter + 1} complete. "
                    f"Waiting {wait}s before run {_iter + 2} of {iterations}..."
                )
                elapsed = 0
                while elapsed < wait:
                    chunk = min(30, wait - elapsed)
                    time.sleep(chunk)
                    elapsed += chunk
                    remaining = wait - elapsed
                    if remaining > 0:
                        self.log.emit(
                            f"[ITERATION] {elapsed}s elapsed, {remaining}s remaining..."
                        )
                self.log.emit(f"[ITERATION] Wait complete. Starting run {_iter + 2}...\n")

# ---------------- Main Window ----------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.run_in_progress = False
        self.setWindowTitle(f"{APP_NAME} {APP_VERSION}")
        self.resize(1200, 820)
        self.setMinimumSize(1200, 820)
        self._init_state()
        self._build_ui()
        # IMPORTANT: fix initial visibility after widgets exist
        QTimer.singleShot(0, self._post_init_layout_fix)

    def _init_state(self):
        self.operation_type = "WLC & AP"
        
        self.workflow = "Custom CLI Commands"
        self.ap_mode = "AP Custom Cmd List"
        self.ap_filter_mode = "NONE"
        self.site_tag = ""
        self.model_group = "All AP Models"
        self.ap_name_map = {}  # NEW: map ip -> ap name for AP Table first column
        self.ap_device = "cos_qca"
        self.wlc_entries = []
        self.ap_list_file = ""
        self.ap_list_path = ""
        self.run_count = 0
        # ── Iterations ──────────────────────────────────────
        self.iterations_enabled = False
        self.iteration_count    = 1
        self.iteration_interval = 300   # seconds
        self.headless_mode = False
        self.enable_debug_collection = False
        self.test_after_iteration = False
        self.wlc_cmds: List[str] = []
        self.ap_cmds: List[str] = []    # multi-WLC list
        if IniStore:
            self.ini = IniStore(CONFIG_FILE)
        else:
            self.ini = None

    def _build_ui(self):
        main = QWidget()
        main_layout = QVBoxLayout(main)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.setCentralWidget(main)

        hero = QWidget()
        hero_layout = QHBoxLayout(hero)
        hero_layout.setContentsMargins(16, 8, 16, 8)
        hero.setStyleSheet("background: #000000;")
        left_logo = QLabel()
        left_pix = safe_pixmap("assets/cisco_logo.png", QSize(100, 36))
        if left_pix:
            left_logo.setPixmap(left_pix)
        hero_layout.addWidget(left_logo, 0, Qt.AlignmentFlag.AlignVCenter)
        title = QLabel("WLAN POLLER GUI")
        title.setObjectName("heroTitle")
        title.setFont(FONT_TITLE)
        title.setStyleSheet("color: white;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hero_layout.addWidget(title, 1)
        right_logo = QLabel()
        right_pix = safe_pixmap("assets/wlc_9800.png", QSize(120, 36))
        if right_pix:
            right_logo.setPixmap(right_pix)
        hero_layout.addWidget(right_logo, 0, Qt.AlignmentFlag.AlignVCenter)
        main_layout.addWidget(hero)

        content = QWidget()
        content_layout = QHBoxLayout(content)
        content_layout.setContentsMargins(14, 28, 14, 14)
        content_layout.setSpacing(12)
        main_layout.addWidget(content, 1)

        self.sidebar = QListWidget()
        self.sidebar.setFixedWidth(220)
        self.sidebar.setFont(QFont("Roboto", 12))
        steps = [
            "Step1 Operation Type",
            "Step2 Credentials",
            "Step3 Workflow",
            "Step4 CLI Cmd List",
            "Step5 AP Filters",
            "Step6 Preview",
            "Step7 Run/Results",
            "Parser",
        ]
        self.sidebar.addItems(steps)
        # self.sidebar.setEnabled(False)
        self.sidebar.setEnabled(True)
        # Make clicking the side nav change pages: connect the currentRowChanged
        # signal to the existing _goto_step method.
        # _goto_step expects an int index, so connect directly.
        self.sidebar.currentRowChanged.connect(self._goto_step)
        content_layout.addWidget(self.sidebar)

        self.stack = QStackedWidget()
        content_layout.addWidget(self.stack, 1)

        self.stack.addWidget(self._page_step1())
        self.stack.addWidget(self._page_step2())
        self.stack.addWidget(self._page_step3())
        self.stack.addWidget(self._page_step4())
        self.stack.addWidget(self._page_step5())
        self.stack.addWidget(self._page_step6())
        self.stack.addWidget(self._page_step7())
        self.stack.addWidget(self._page_parser())

        self._goto_step(0)
        QTimer.singleShot(0, self._refresh_visibility)

    def _stop_worker(self, timeout_ms: int = 3000):
        """
        Try to stop self.worker cleanly. If it doesn't stop within timeout, terminate it.
        Called during app close to avoid hanging background threads.
        """
        try:
            if not hasattr(self, "worker") or self.worker is None:
                return
            w = self.worker
            # Request interruption if QThread supports it
            try:
                w.requestInterruption()
            except Exception:
                pass
            # Ask QThread to quit (if run() listens for interruption this will help)
            try:
                w.quit()
            except Exception:
                pass
            # Wait for completion briefly
            try:
                w.wait(timeout_ms)
            except Exception:
                pass
            # If still running, force terminate (last resort)
            if getattr(w, "isRunning", lambda: False)():
                try:
                    w.terminate()
                except Exception:
                    pass
        except Exception:
            pass

    def closeEvent(self, event):
        if getattr(self, "run_in_progress", False):
            QMessageBox.warning(
                self,
                "Operation Running",
                "Please wait until the operation completes."
            )
            event.ignore()
            return

        event.accept()
    
    def _inject_run_preview_into_log(self):
        """
        Inject Step6 preview block at top of Step7 Run Log.
        Production safe: no passwords exposed.
        """

        if not hasattr(self, "run_log"):
            return

        preview_text = ""

        # Get Step6 preview text safely
        if hasattr(self, "preview_text"):
            try:
                preview_text = self.preview_text.toPlainText().strip()
            except Exception:
                preview_text = ""

        if not preview_text:
            preview_text = "Preview not available."

        header_block = []
        header_block.append("=" * 56)
        header_block.append("RUN CONFIGURATION PREVIEW")
        header_block.append("=" * 56)
        header_block.append(preview_text.strip())
        header_block.append("=" * 56)
        header_block.append("STARTING EXECUTION...")
        header_block.append("=" * 56)
        header_block.append("")

        try:
            self.run_log.append("\n".join(header_block))
        except Exception:
            pass
    def _load_and_run_cronjob(self):
        """Load CRONJOB config + load commands from files (NOT config.ini) and auto-run."""

        if not self.ini or not self.ini.cfg.has_option("CRONJOB", "operation_type"):
            return

        g = lambda k, d="": self.ini.cfg.get("CRONJOB", k, fallback=d)

        # ---------------- BASIC CONFIG ----------------
        self.operation_type = g("operation_type", "WLC & AP")
        self.workflow        = g("workflow", "Custom CLI Commands")
        self.ap_mode         = g("ap_mode", "AP Custom Cmd List")
        self.ap_filter_mode  = g("ap_filter_mode", "NONE")
        self.site_tag        = g("site_tag", "")
        self.model_group     = g("model_group", "All AP Models")
        self.ap_device              = g("ap_device", "cos_qca")
        self.enable_debug_collection = False  # now controlled by counter, not UI toggle
        # ── Iteration state ──────────────────────────────────────
        self.iterations_enabled  = g("iterations_enabled", "false").lower() == "true"
        try:
            self.iteration_count = max(1, min(50, int(g("iteration_count", "1"))))
        except Exception:
            self.iteration_count = 1
        try:
            self.iteration_interval = max(0, min(18000, int(g("iteration_interval", "300"))))
        except Exception:
            self.iteration_interval = 300

        # Safety: Client Stuck In Auth Loop is always WLC Only — fix mis-saved cronjob
        if self.workflow == "Client Stuck In Auth Loop":
            self.operation_type = "WLC Only"
       # ---------------- AP LIST FILE RESOLUTION ----------------
        # ---------------- AP LIST FILE RESOLUTION (FIXED) ----------------
        self.ap_list_file = g("ap_list_file", "") or \
                            self.ini.cfg.get("GENERAL", "last_ap_list_file", fallback="")

        if not self.ap_list_file:
            print("[CRONJOB] ERROR: AP list file not set")

        else:
            # Convert to absolute path
            if not os.path.isabs(self.ap_list_file):
                self.ap_list_file = os.path.join(BASE_DIR, self.ap_list_file)

            # ✅ PRIMARY CHECK
            if not os.path.exists(self.ap_list_file):

                # ✅ FALLBACK TO ENGINE LOCATION
                fallback = os.path.join(CONFD_DIR, "ap_ip_list.txt")

                if os.path.exists(fallback):
                    print(f"[CRONJOB] Using fallback AP list: {fallback}")
                    self.ap_list_file = fallback

                else:
                    print(f"[CRONJOB] ERROR: AP list not found anywhere")
                    print(f"[CRONJOB] Tried: {self.ap_list_file}")
                    print(f"[CRONJOB] Tried fallback: {fallback}")

            print(f"[CRONJOB] FINAL AP list file: {self.ap_list_file}")
        # ---------------- LOAD COMMANDS FROM FILES ----------------
        def _load_cmds(path):
            if not os.path.exists(path):
                print(f"[CRONJOB] File not found: {path}")
                return []
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return [line.strip() for line in f if line.strip()]

        wlc_file = os.path.join(CONFD_DIR, "wlc_cmds.txt")
        ap_file  = os.path.join(CONFD_DIR, "ap_cmds.txt")

        self.wlc_cmds = _load_cmds(wlc_file)
        self.ap_cmds  = _load_cmds(ap_file)

        print(f"[CRONJOB] Loaded commands from files")
        print(f"[CRONJOB] WLC cmds count = {len(self.wlc_cmds)}")
        print(f"[CRONJOB] AP cmds count  = {len(self.ap_cmds)}")

        # ---------------- VALIDATION ----------------
        

        # ---------------- SYNC UI ----------------
        try:
            if hasattr(self, "wlc_cmd_box"):
                self.wlc_cmd_box.setPlainText("\n".join(self.wlc_cmds))
        except Exception:
            pass

        try:
            if hasattr(self, "ap_cmd_box"):
                self.ap_cmd_box.setPlainText("\n".join(self.ap_cmds))
        except Exception:
            pass

        # ---------------- SYNC OPERATION DROPDOWN ----------------
        try:
            idx = {"WLC Only": 0, "WLC & AP": 1, "AP Only": 2}.get(self.operation_type, 1)
            self.op_dd.blockSignals(True)
            self.op_dd.setCurrentIndex(idx)
            self.op_dd.blockSignals(False)
        except Exception:
            pass

        # ---------------- LOAD WLC CREDENTIALS ----------------
        try:
            wlc_secs = []
            for sec in self.ini.cfg.sections():
                if sec == "WLC" or (sec.startswith("WLC") and sec[3:].isdigit()):
                    if self.ini.get(sec, "wlc_ip"):
                        wlc_secs.append(sec)

            wlc_secs.sort(key=lambda s: 0 if s == "WLC" else int(s[3:]))

            # Clear UI
            while self.wlc_entries:
                e = self.wlc_entries.pop()
                try:
                    e["widget"].deleteLater()
                except Exception:
                    pass

            # Rebuild UI
            for _ in wlc_secs:
                self._add_wlc_entry()

            for i, sec in enumerate(wlc_secs):
                if i < len(self.wlc_entries):
                    self.wlc_entries[i]["ip"].setText(self.ini.get(sec, "wlc_ip"))
                    self.wlc_entries[i]["user"].setText(self.ini.get(sec, "wlc_user"))
                    self.wlc_entries[i]["pasw"].setText(self.ini.get(sec, "wlc_pasw"))

        except Exception as e:
            print(f"[CRONJOB] WLC rebuild failed: {e}")

        # ---------------- LOAD AP CREDS ----------------
        try:
            if self.operation_type in ("WLC & AP", "AP Only"):
                if hasattr(self, "ap_user"):
                    self.ap_user.setText(self.ini.get("AP", "ap_user"))
                if hasattr(self, "ap_pass"):
                    self.ap_pass.setText(self.ini.get("AP", "ap_pasw"))
                if hasattr(self, "ap_enable"):
                    self.ap_enable.setText(self.ini.get("AP", "ap_enable"))
        except Exception:
            pass

        # ---------------- WORKFLOW DROPDOWN ----------------
        try:
            saved_wf = self.workflow
            self._update_workflow_dropdown()

            idx = self.workflow_dd.findText(saved_wf)
            if idx >= 0:
                self.workflow_dd.blockSignals(True)
                self.workflow_dd.setCurrentIndex(idx)
                self.workflow_dd.blockSignals(False)

            self.workflow = saved_wf
        except Exception:
            pass

        # ---------------- START RUN ----------------
        print("[CRONJOB] Auto-starting run...")
        QTimer.singleShot(500, self._start_run)
    def _build_per_wlc_cmd_boxes(self):

        layout = self.per_wlc_cmd_section.layout()

        # clear old
        for i in reversed(range(layout.count())):
            item = layout.itemAt(i)
            w = item.widget()
            if w:
                w.deleteLater()

        self.per_wlc_cmd_boxes = {}

        for i, entry in enumerate(self.wlc_entries):

            ip = entry["ip"].text()

            group = QGroupBox(f"WLC {i+1}")
            g_layout = QVBoxLayout(group)

            ip_label = QLabel(f"IP: {ip}")
            ip_label.setStyleSheet("color:#6b7280; font-size:11px;")

            cmd_box = QTextEdit()
            cmd_box.setPlaceholderText("Enter commands (one per line)")
            cmd_box.setFixedHeight(120)

            g_layout.addWidget(ip_label)
            g_layout.addWidget(cmd_box)

            layout.addWidget(group)

            self.per_wlc_cmd_boxes[f"WLC{i+1}"] = cmd_box

            # load saved
            section_name = f"WLC{i+1}"
            if self.ini and self.ini.cfg.has_section(f"{section_name}_CMDS"):
                raw = self.ini.cfg.get(f"{section_name}_CMDS", "cmds", fallback="")
                cmd_box.setPlainText(raw)
            
    # ---------------- Pages ----------------
    def _page_step1(self) -> QWidget:
        """
        Step1 page: card with Choose Operation Type + AP upload (card),
        and the Enter Credentials button placed OUTSIDE the card (below it),
        matching Step2 layout.
        """
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)
        # lay.setContentsMargins(12, 12, 12, 12)
        lay.setAlignment(Qt.AlignTop)

        # --- Card ---
        card = QGroupBox("Step1 - Select Operation Type")
        card.setFont(FONT_CARD_TITLE)
        # Make the card only as tall as its contents (prevents it from filling the page)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

        c_l = QVBoxLayout(card)
        # c_l.setContentsMargins(12, 12, 12, 12)
        c_l.setContentsMargins(25, 25, 25, 25)
        c_l.setSpacing(18)

        # Label + combobox
        lbl = QLabel("Choose Operation Type:")
        lbl.setStyleSheet("padding-top:4px;")
        c_l.addWidget(lbl)

        self.op_dd = QComboBox()
        # keep same 3 choices but default to "WLC & AP"
        self.op_dd.addItems(["WLC Only", "WLC & AP", "AP Only"])
        self.op_dd.currentTextChanged.connect(self._on_operation_change)
        self.op_dd.setCurrentIndex(1)  # index 1 -> "WLC & AP"
        self.op_dd.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.op_dd.setFixedHeight(30)
        c_l.addWidget(self.op_dd)

        # AP upload row (still inside card)
        self.ap_upload_row = QWidget()
        ab = QHBoxLayout(self.ap_upload_row)
        ab.setContentsMargins(0, 0, 0, 0)
        ab.setSpacing(8)
        ab.addWidget(QLabel("Upload AP List File (Format: AP Ip, AP Name)"))

        self.ap_path = QLineEdit()
        self.ap_path.setReadOnly(True)
        ab.addWidget(self.ap_path)
        self.ap_stats = QLabel("")
        self.ap_stats.setStyleSheet("color:#374151; font-weight:600;")


        self.ap_browse = QPushButton("Browse")
        self.ap_browse.setProperty("class", "secondary")
        self.ap_browse.clicked.connect(self._browse_ap_list)
        ab.addWidget(self.ap_browse)

        c_l.addWidget(self.ap_upload_row)
        self.ap_stats = QLabel("")
        self.ap_stats.setAlignment(Qt.AlignLeft)

        self.ap_stats.setStyleSheet("""
        font-weight:600;
        padding-top:4px;
        """)

        c_l.addWidget(self.ap_stats)

        # small breathing room inside the card (no large stretch)
        c_l.addSpacing(6)

        # --- add card to page layout ---
        lay.addWidget(card)

        # --- Controls row (OUTSIDE the card, just like Step2) ---
        controls = QHBoxLayout()
        controls.setContentsMargins(6, 6, 6, 6)
        controls.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        c_l.addSpacing(6)

        self.btn_step1_next = QPushButton("Enter Credentials")
        self.btn_step1_next.setProperty("nav", True)  # keep nav styling if desired
        self.btn_step1_next.setFixedHeight(34)
        self.btn_step1_next.clicked.connect(self._step1_enter_credentials)
        controls.addWidget(self.btn_step1_next)

        # Add controls *below* the card (not inside it)
        lay.addLayout(controls)

        # Refresh state and return widget
        self._refresh_step1()
        return w
    def _step1_enter_credentials(self):
        if self.op_dd.currentText() == "AP Only":
            if not self.ap_list_file or not os.path.isfile(self.ap_list_file):
                QMessageBox.warning(
                    self,
                    "AP List Required",
                    "Please upload an AP list file to proceed."
                )
                return

        self._goto_step(1)

        
        
    def _page_step2(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        card = QGroupBox("Step2 - WLC / AP Details")
        card.setFont(FONT_CARD_TITLE)
        card_v = QVBoxLayout(card)
        card_v.setContentsMargins(12, 12, 12, 12)
        card_v.setSpacing(6)
        self.step2_top_spacer = QSpacerItem(0, 15, QSizePolicy.Minimum, QSizePolicy.Fixed)
        card_v.addItem(self.step2_top_spacer)
        card_v.addSpacing(15)
        # ---------- WLC BLOCK ----------
        self.wlc_block = QWidget()
        wlc_outer = QVBoxLayout(self.wlc_block)
        wlc_outer.setContentsMargins(0, 0, 0, 0)
        wlc_outer.setSpacing(6)

        # 🔘 Configuration mode row (Manual vs. Bulk Upload) — own row, above the
        # WLC Configurations header, so it reads as a mode switch rather than a
        # toolbar toggle.
        wlc_mode_row = QHBoxLayout()
        wlc_mode_row.addWidget(QLabel("Configuration Mode:"))

        self.wlc_mode_manual_radio = QRadioButton("Manual Entry (up to 3 WLC's)")
        self.wlc_mode_bulk_radio = QRadioButton("Bulk Upload (up to 100 WLC's)")
        self.wlc_mode_manual_radio.setChecked(True)

        self.wlc_mode_group = QButtonGroup(self.wlc_block)
        self.wlc_mode_group.addButton(self.wlc_mode_manual_radio)
        self.wlc_mode_group.addButton(self.wlc_mode_bulk_radio)
        self.wlc_mode_bulk_radio.toggled.connect(self._on_multi_wlc_toggle)

        wlc_mode_row.addWidget(self.wlc_mode_manual_radio)
        wlc_mode_row.addWidget(self.wlc_mode_bulk_radio)
        wlc_mode_row.addStretch()
        wlc_outer.addLayout(wlc_mode_row)

        wlc_header = QHBoxLayout()
        wlc_header.addWidget(QLabel("WLC Configurations"))

        # ➖ Remove WLC (NEW)
        self.remove_btn_wlc = QPushButton("Remove WLC")
        self.remove_btn_wlc.setFixedHeight(28)
        self.remove_btn_wlc.clicked.connect(self._remove_last_wlc_entry)

        # ➕ Add WLC
        self.btn_add_wlc = QPushButton("Add WLC")
        self.btn_add_wlc.setFixedHeight(28)
        self.btn_add_wlc.clicked.connect(self._add_wlc_entry)

        wlc_header.addStretch()
        wlc_header.addWidget(self.remove_btn_wlc)
        wlc_header.addWidget(self.btn_add_wlc)
        wlc_outer.addLayout(wlc_header)

        # 📂 Bulk WLC list upload row (Excel) — shown only in Bulk Upload mode
        self.wlc_bulk_upload_row = QWidget()
        wb = QHBoxLayout(self.wlc_bulk_upload_row)
        wb.setContentsMargins(0, 0, 0, 0)
        wb.setSpacing(8)
        wb.addWidget(QLabel("Upload WLC List (Excel: WLC IP, WLC Name)"))

        self.wlc_bulk_path = QLineEdit()
        self.wlc_bulk_path.setReadOnly(True)
        wb.addWidget(self.wlc_bulk_path)

        self.wlc_bulk_browse = QPushButton("Browse")
        self.wlc_bulk_browse.setProperty("class", "secondary")
        self.wlc_bulk_browse.clicked.connect(self._browse_wlc_list_excel)
        wb.addWidget(self.wlc_bulk_browse)

        self.wlc_bulk_upload_row.setVisible(False)
        wlc_outer.addWidget(self.wlc_bulk_upload_row)

        # Validation/status feedback for the uploaded file (mirrors ap_stats).
        # Populated once the Excel parsing backend is wired up.
        self.wlc_bulk_stats = QLabel("")
        self.wlc_bulk_stats.setStyleSheet("color:#374151; font-weight:600; font-size:11px;")
        self.wlc_bulk_stats.setVisible(False)
        wlc_outer.addWidget(self.wlc_bulk_stats)

        self.wlc_entries_widget = QWidget()
        self.wlc_entries_layout = QVBoxLayout(self.wlc_entries_widget)
        self.wlc_entries_layout.setContentsMargins(0, 0, 0, 0)
        self.wlc_entries_layout.setSpacing(8)
        wlc_outer.addWidget(self.wlc_entries_widget)

        # ℹ️ Guidance callout for bulk WLC mode — shown only in Bulk Upload mode
        self.wlc_multi_note = QLabel(
            "<b>⚠ All WLCs must use the same credentials.</b><br>"
            "All WLC connections authenticate using the same username and password "
            "entered here — confirm they are valid on every WLC before proceeding."
            "<br><span style='font-size:10px;'>The uploaded file must be in Excel "
            "format and list each WLC's IP address (required) and name (optional)."
            "</span>"
        )
        self.wlc_multi_note.setWordWrap(True)
        self.wlc_multi_note.setStyleSheet(
            "background:#fffbeb; border:1px solid #fcd34d; border-radius:6px; "
            "color:#92400e; font-size:11px; padding:8px 10px;"
        )
        self.wlc_multi_note.setVisible(False)
        wlc_outer.addWidget(self.wlc_multi_note)

        card_v.addWidget(self.wlc_block)
        self._add_wlc_entry()   # seed first entry

        # ---------- AP BLOCK ----------
        self.ap_block = QWidget()
        ap_form = QFormLayout(self.ap_block)

        self.ap_user_label = QLabel("AP Username")
        self.ap_user = QLineEdit(self.ini.get("AP", "ap_user") if self.ini else "")
        self.ap_user.setFixedHeight(30)

        self.ap_pass_label = QLabel("AP Password")
        self.ap_pass = QLineEdit(self.ini.get("AP", "ap_pasw") if self.ini else "")
        self.ap_pass.setEchoMode(QLineEdit.Password)
        self.ap_pass.setFixedHeight(30)

        self.ap_enable_label = QLabel("Enable Password")
        self.ap_enable = QLineEdit(self.ini.get("AP", "ap_enable") if self.ini else "")
        self.ap_enable.setEchoMode(QLineEdit.Password)
        self.ap_enable.setFixedHeight(30)

        ap_form.addRow(self.ap_user_label, self.ap_user)
        ap_form.addRow(self.ap_pass_label, self.ap_pass)
        ap_form.addRow(self.ap_enable_label, self.ap_enable)

        card_v.addWidget(self.ap_block)

        lay.addWidget(card)

        row = QHBoxLayout()
        back_btn = QPushButton("Back")
        back_btn.setProperty("nav", True)
        back_btn.clicked.connect(lambda: self._goto_step(0))

        save_btn = QPushButton("Save")
        save_btn.setProperty("nav", True)
        save_btn.clicked.connect(self._save_creds)

        proceed_btn = QPushButton("Proceed")
        proceed_btn.setProperty("nav", True)
        proceed_btn.clicked.connect(self._step2_proceed)

        row.addWidget(back_btn);
        row.addWidget(save_btn);
        row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum));
        row.addWidget(proceed_btn)
        lay.addLayout(row)
        self._on_operation_change(self.operation_type)
        return w

    def _page_step3(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        # Card
        card = QGroupBox("Step3 - Choose WorkFlow")
        card.setFont(FONT_CARD_TITLE)
        # Keep card only as tall as its contents
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

        c_l = QVBoxLayout(card)
        c_l.setContentsMargins(25, 25, 25, 25)
        c_l.setSpacing(25)
        # Label
        c_l.addWidget(QLabel("Choose a WorkFlow"))

        # ── WORKFLOW DROPDOWN ─────────────────────────
        self.workflow_dd = QComboBox()
        self.workflow_dd.addItems([
            "Custom CLI Commands",
            "AP Flash Checker",
            "Upload Files from AP",   # if already exists keep it
            "AP Cleanup + Reload"     # ✅ NEW WORKFLOW
        ])
        self.workflow_dd.currentTextChanged.connect(self._on_workflow_change)
        self.workflow_dd.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.workflow_dd.setFixedHeight(30)
        c_l.addWidget(self.workflow_dd)
        

       
       
        # ── UPLOAD CONFIG ─────────────────────────────
        self.upload_config_widget = QWidget()
        self.upload_config_widget.setStyleSheet(
            "background:#f0f9ff; border:1px solid #bae6fd; border-radius:6px; padding:4px;"
        )

        upload_cfg_layout = QFormLayout(self.upload_config_widget)
        upload_cfg_layout.setContentsMargins(12, 10, 12, 10)
        upload_cfg_layout.setSpacing(8)

        upload_title = QLabel("Upload Settings")
        upload_title.setStyleSheet("font-weight:600; color:#0369a1; font-size:12px;")
        upload_cfg_layout.addRow(upload_title)

        self.upload_file_type_dd = QComboBox()
        self.upload_file_type_dd.addItems(["SupportBundle"])
        upload_cfg_layout.addRow("File Type:", self.upload_file_type_dd)

        self.upload_proto_dd = QComboBox()
        self.upload_proto_dd.addItems(["TFTP", "SFTP"])
        self.upload_proto_dd.currentTextChanged.connect(self._on_upload_proto_changed)
        upload_cfg_layout.addRow("Protocol:", self.upload_proto_dd)

        self.upload_server_ip_field = QLineEdit()
        self.upload_server_ip_field.setPlaceholderText("Server IP  e.g. 192.168.0.10")
        upload_cfg_layout.addRow("Server IP:", self.upload_server_ip_field)

        self.upload_sftp_user_label = QLabel("SFTP Username:")
        self.upload_sftp_user = QLineEdit()
        self.upload_sftp_user.setVisible(False)
        self.upload_sftp_user_label.setVisible(False)

        self.upload_sftp_pass_label = QLabel("SFTP Password:")
        self.upload_sftp_pass = QLineEdit()
        self.upload_sftp_pass.setEchoMode(QLineEdit.Password)
        self.upload_sftp_pass.setVisible(False)
        self.upload_sftp_pass_label.setVisible(False)

        upload_cfg_layout.addRow(self.upload_sftp_user_label, self.upload_sftp_user)
        upload_cfg_layout.addRow(self.upload_sftp_pass_label, self.upload_sftp_pass)

        self.upload_config_widget.setVisible(False)

        c_l.addWidget(self.upload_config_widget)

        

        c_l.addSpacing(6)

        # Add the card to page layout
        lay.addWidget(card)

        # Controls row (OUTSIDE the card, same style as Step1/Step2)
        row = QHBoxLayout()
        row.setContentsMargins(6, 6, 6, 6)
        row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))

        back = QPushButton("Back")
        back.setProperty("nav", True)
        back.clicked.connect(lambda: self._goto_step(1))  # go back to Step2 Credentials
        row.addWidget(back)

        nextb = QPushButton("Proceed")
        nextb.setProperty("nav", True)
        nextb.clicked.connect(self._step3_proceed)
        row.addWidget(nextb)

        # Align the row the same way as in Step1/Step2: right-aligned
        # (we already added an expanding spacer before the back button)
        lay.addLayout(row)
        

        
       
        return w

    def _page_step4(self) -> QWidget:
        from PySide6.QtWidgets import QScrollArea

        outer = QWidget()
        outer_lay = QVBoxLayout(outer)
        outer_lay.setContentsMargins(0, 0, 0, 0)
        outer_lay.setSpacing(4)

        # ── SCROLL AREA wrapping the card ──
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        card = QGroupBox("Step4 - CLI Cmd List")
        card.setFont(FONT_CARD_TITLE)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        c_l = QVBoxLayout(card)
        c_l.setContentsMargins(12, 18, 12, 12)
        c_l.setSpacing(10)

        # ── WLC CMD BOX ──────────────────────────────────────
        self.wlc_cmd_box = QTextEdit()
        self.wlc_cmd_box.setPlaceholderText("Enter WLC commands (one per line)")
        self.wlc_cmd_box.setFixedHeight(160)
        self.wlc_cmd_box.setAutoFillBackground(True)
        self.wlc_cmd_box.setStyleSheet(
            "QTextEdit { background-color: #ffffff !important; border: 1px solid #e6e8eb; border-radius: 6px; padding: 6px; }")
        self.wlc_cmd_label = QLabel("WLC Cmd List:")
        self.wlc_cmd_label.setStyleSheet("font-weight:600;")
        self.wlc_cmd_section = QWidget()
        self.wlc_cmd_section.setStyleSheet("background: #ffffff;")
        self.wlc_cmd_section.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        wlc_layout = QVBoxLayout(self.wlc_cmd_section)
        wlc_layout.setContentsMargins(0, 0, 0, 0)
        wlc_layout.setSpacing(6)
        wlc_layout.addWidget(self.wlc_cmd_label)
        wlc_layout.addWidget(self.wlc_cmd_box)
        c_l.addWidget(self.wlc_cmd_section)

        # ── AP SECTION ───────────────────────────────────────
        self.ap_cmd_section = QWidget()
        self.ap_cmd_section.setStyleSheet("background: #ffffff;")
        self.ap_cmd_section.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        ap_section_layout = QVBoxLayout(self.ap_cmd_section)
        ap_section_layout.setContentsMargins(0, 0, 0, 0)
        ap_section_layout.setSpacing(8)

        ap_cmd_label = QLabel("AP Cmd List:")
        ap_cmd_label.setStyleSheet("font-weight: bold; margin-top: 4px;")
        ap_section_layout.addWidget(ap_cmd_label)

        ap_mode_row = QHBoxLayout()
        ap_mode_row.addWidget(QLabel("AP Mode:"))
        self.ap_mode_dd = QComboBox()
        self.ap_mode_dd.addItems(["AP Custom Cmd List", "AP Image Download"])
        self.ap_mode_dd.currentTextChanged.connect(self._on_ap_mode_changed)
        ap_mode_row.addWidget(self.ap_mode_dd)
        ap_mode_row.addStretch()
        ap_section_layout.addLayout(ap_mode_row)
# ── AP DATAPATH QUEUE MON CONFIG (only for this workflow) ──
        self.datapath_mon_widget = QWidget()
        dp_form = QFormLayout(self.datapath_mon_widget)
        dp_form.setContentsMargins(0, 4, 0, 4)
        dp_title = QLabel("AP Datapath Queue Monitor")
        dp_title.setStyleSheet("font-weight:600; color:#0369a1;")
        dp_form.addRow(dp_title)
        dp_note = QLabel(
            f"Runs against ALL clients/Radios/VAPs discovered per-AP via "
            f"'show client summary' — no MAC entry needed. "
            f"Collection runs {DATAPATH_MON_DEFAULT_ITERATIONS} iterations, "
            f"{DATAPATH_MON_DEFAULT_INTERVAL_SEC}s apart (fixed in code)."
        )
        dp_note.setWordWrap(True)
        dp_note.setStyleSheet("color:#6b7280; font-size:11px;")
        dp_form.addRow(dp_note)
        self.datapath_mon_widget.setVisible(False)
        ap_section_layout.addWidget(self.datapath_mon_widget)
        self.ap_cmd_box = QTextEdit()
        self.ap_cmd_box.setPlaceholderText("Enter AP CLI commands (one per line)")
        self.ap_cmd_box.setFixedHeight(160)
        self.ap_cmd_box.setStyleSheet(
            "QTextEdit { background: #ffffff; border: 1px solid #e6e8eb; border-radius: 6px; }")
        ap_section_layout.addWidget(self.ap_cmd_box)

        # ── IMAGE DOWNLOAD SETTINGS ──────────────────────────
        self.ftp_group = QGroupBox("AP Image Download Settings")
        self.ftp_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        ftp_layout = QFormLayout()
        ftp_layout.setSpacing(12)
        ftp_layout.setContentsMargins(12, 16, 12, 16)
        self.ftp_group.setLayout(ftp_layout)

        self.proto_dd = QComboBox()
        self.proto_dd.addItems(["TFTP", "SFTP"])
        self.proto_dd.setFixedHeight(30)
        self.proto_dd.currentTextChanged.connect(self._on_proto_changed)
        ftp_layout.addRow("Protocol:", self.proto_dd)

        self.ftp_user_label = QLabel("SFTP Username:")
        self.ftp_user = QLineEdit()
        self.ftp_user.setPlaceholderText("SFTP username")
        self.ftp_user.setFixedHeight(30)
        self.ftp_user.setVisible(False)
        self.ftp_user_label.setVisible(False)

        self.ftp_pasw_label = QLabel("SFTP Password:")
        self.ftp_pasw = QLineEdit()
        self.ftp_pasw.setPlaceholderText("SFTP password")
        self.ftp_pasw.setEchoMode(QLineEdit.Password)
        self.ftp_pasw.setFixedHeight(30)
        self.ftp_pasw.setVisible(False)
        self.ftp_pasw_label.setVisible(False)

        ftp_layout.addRow(self.ftp_user_label, self.ftp_user)
        ftp_layout.addRow(self.ftp_pasw_label, self.ftp_pasw)

        self.ftp_addr = QLineEdit()
        self.ftp_path = QLineEdit()

        self.ftp_group.setVisible(False)
        ap_section_layout.addWidget(self.ftp_group)

        c_l.addWidget(self.ap_cmd_section)
        lay.addWidget(card)
        lay.addStretch()

        scroll.setWidget(w)
        outer_lay.addWidget(scroll, 1)

        # ── BUTTONS (outside scroll, always visible) ─────────
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(8, 6, 8, 12)
        outer_lay.setContentsMargins(0, 0, 0, 0)

        back_btn = QPushButton("Back")
        back_btn.setProperty("nav", True)
        back_btn.clicked.connect(lambda: self._goto_step(2))
        

        save_btn = QPushButton("Save")
        save_btn.setProperty("nav", True)
        save_btn.clicked.connect(self._step4_save)

        proceed_btn = QPushButton("Proceed")
        proceed_btn.setProperty("nav", True)
        proceed_btn.clicked.connect(self._step4_proceed)

        btn_row.addWidget(back_btn)
        btn_row.addWidget(save_btn)
        btn_row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        btn_row.addWidget(proceed_btn)
        outer_lay.addLayout(btn_row)

        self._on_ap_mode_changed(self.ap_mode_dd.currentText())
        return outer
    
            
    def _page_step5(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(6)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        card = QGroupBox("Step5 - AP Filters")
        card.setFont(FONT_CARD_TITLE)

        c_l = QGridLayout()
        c_l.setContentsMargins(12, 12, 12, 12)
        c_l.setHorizontalSpacing(12)
        c_l.setVerticalSpacing(6)
        c_l.setSpacing(15)

        r = 0
        # note = QLabel("(Only one filter can be active at a time)")
        # note.setStyleSheet(f"color:{TEXT_MUTED};")
        # c_l.addWidget(note, r, 0, 1, 2)
        r += 1

        # --- APs By Model (show first) ---
        self.chk_model = QCheckBox("APs By Model")
        self.model_dd = QComboBox()
        self.model_dd.addItems([
            "All AP Models",
            "AP1852/2802/3802/4802",
            "C9105AX/9115AX/9120AX",
            "C9117AX/9130AX/9136/9124",
            "C9162/9163/9164/9166",
            "C9171/9172/9174/9176/9178/9179"
        ])
        self.model_dd.setFixedHeight(28)
        self.model_dd.setEnabled(False)
        self.chk_model.toggled.connect(lambda on: self.model_dd.setEnabled(on))
        c_l.addWidget(self.chk_model, r, 0)
        c_l.addWidget(self.model_dd, r, 1)
        r += 1

        # separator
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        c_l.addWidget(sep, r, 0, 1, 2)
        r += 1

        # --- APs By SiteTag (show second) ---
        self.chk_site = QCheckBox("APs By SiteTag(optional)")
        self.site_tag_txt = QLineEdit()
        self.site_tag_txt.setPlaceholderText("Enter SiteTag Name")
        self.site_tag_txt.setFixedHeight(28)
        self.site_tag_txt.setEnabled(False)
        self.chk_site.toggled.connect(lambda on: self.site_tag_txt.setEnabled(on))
        c_l.addWidget(self.chk_site, r, 0)
        c_l.addWidget(self.site_tag_txt, r, 1)
        r += 1

        # Ensure only one filter active at a time
        self.chk_site.toggled.connect(self._enforce_one_filter)
        self.chk_model.toggled.connect(self._enforce_one_filter)

        card.setLayout(c_l)
        lay.addWidget(card)

        # Buttons row
        row = QHBoxLayout()
        back = QPushButton("Back")
        back.setProperty("nav", True)
        back.clicked.connect(lambda: self._goto_step(3 if self.workflow == "Custom CLI Commands" else 2))

        prev = QPushButton("Preview")
        prev.setProperty("nav", True)
        prev.clicked.connect(self._step5_preview)

        row.addWidget(back)
        row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        row.addWidget(prev)
        lay.addLayout(row)
        lay.addStretch()

        return w

    def _page_step6(self) -> QWidget:
        w = QWidget()

        lay = QVBoxLayout(w)
        lay.setSpacing(12)
        lay.setContentsMargins(12, 12, 12, 12)

        # -------- CARD --------
        card = QGroupBox("Step6 - Preview")
        card.setFont(FONT_CARD_TITLE)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        c_l = QVBoxLayout(card)
        c_l.setContentsMargins(18, 18, 18, 18)

        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        c_l.addWidget(self.preview_text)

        # ── ITERATION CONFIG as collapsible dropdown ─────────────
        iter_toggle_btn = QPushButton("▶  Iteration Config (optional)")
        iter_toggle_btn.setCheckable(True)
        iter_toggle_btn.setChecked(False)
        iter_toggle_btn.setStyleSheet("""
            QPushButton {
                text-align: left;
                padding: 6px 12px;
                background: #f3f4f6;
                color: #374151;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                font-weight: 600;
            }
            QPushButton:checked {
                background: #e5e7eb;
            }
        """)

        iter_body = QWidget()
        iter_body.setVisible(False)
        iter_body_layout = QGridLayout(iter_body)
        iter_body_layout.setContentsMargins(12, 10, 12, 10)
        iter_body_layout.setHorizontalSpacing(14)
        iter_body_layout.setVerticalSpacing(6)

        self.chk_iterations = QCheckBox("Enable Iterations")
        self.chk_iterations.setChecked(self.iterations_enabled)
        iter_body_layout.addWidget(self.chk_iterations, 0, 0, 1, 2)

        iter_body_layout.addWidget(QLabel("Count (max 50):"), 1, 0)
        self.iter_count_field = QLineEdit()
        self.iter_count_field.setFixedWidth(80)
        self.iter_count_field.setPlaceholderText("0–50")
        self.iter_count_field.setEnabled(self.iterations_enabled)
        iter_body_layout.addWidget(self.iter_count_field, 1, 1)

        iter_body_layout.addWidget(QLabel("Interval in seconds max(18000):"), 2, 0)
        self.iter_interval_field = QLineEdit()
        self.iter_interval_field.setFixedWidth(80)
        self.iter_interval_field.setPlaceholderText("1–18000")
        self.iter_interval_field.setEnabled(self.iterations_enabled)
        iter_body_layout.addWidget(self.iter_interval_field, 2, 1)

        def _on_iter_toggle(checked):
            iter_body.setVisible(checked)
            iter_toggle_btn.setText(
                ("▼" if checked else "▶") + "  Iteration Config (optional)"
            )

        iter_toggle_btn.toggled.connect(_on_iter_toggle)

        def _on_iter_chk(state):
            on = bool(state)
            self.iter_count_field.setEnabled(on)
            self.iter_interval_field.setEnabled(on)
            self._sync_iter_state()
            self._fill_preview()

        def _sync_and_preview():
            self._sync_iter_state()
            self._fill_preview()

        self.chk_iterations.stateChanged.connect(_on_iter_chk)
        self.iter_count_field.textChanged.connect(lambda _: _sync_and_preview())
        self.iter_interval_field.textChanged.connect(lambda _: _sync_and_preview())
        # ── Static compatibility notice ───────────────────────
        iter_info = QLabel(
            "<b>Not all workflows support iterations.</b><br>"
            "The following will automatically disable this feature:<br>"
            "<ul style='margin:4px 0 0 16px; padding:0;'>"
            "<li><b>AP Image Download</b> — one-time flash operation; "
            "re-running risks storage corruption</li>"
            "<li><b>TMP Cleanup + Reload</b> — APs reboot after this command; "
            "they will be unreachable during the next iteration window</li>"
            
            "and is a one-shot recovery workflow</li>"
            "<li>Any command containing: <code>reload</code>, "
            "<code>archive download-sw</code>, <code>sftp://</code>, "
            "<code>scp://</code></li>"
            "</ul>"
        )
        iter_info.setWordWrap(True)
        iter_info.setTextFormat(Qt.RichText)
        iter_info.setStyleSheet(
            "background: #f0f9ff;"
            "border: 1px solid #bae6fd;"
            "border-radius: 6px;"
            "padding: 8px 12px;"
            "font-size: 11px;"
            "color: #0c4a6e;"
        )
        iter_body_layout.addWidget(iter_info, 3, 0, 1, 2)
        
        # ── END notice ───────────────────────────────────────
        # ── TEST KNOB: Test After Iteration (small, testing-only) ───
        self.chk_test_after_iteration = QCheckBox("Test mode")
        self.chk_test_after_iteration.setChecked(getattr(self, "test_after_iteration", False))
        self.chk_test_after_iteration.setStyleSheet("font-size:10px; color:#9ca3af;")
        self.chk_test_after_iteration.stateChanged.connect(
            lambda state: setattr(self, "test_after_iteration", bool(state))
        )
        iter_body_layout.addWidget(self.chk_test_after_iteration, 4, 0, 1, 2)
        # ── END TEST KNOB ─────────────────────────────────────

        

        c_l.addWidget(iter_toggle_btn)
        c_l.addWidget(iter_body)
        # ── END ITERATION CONFIG ─────────────────────────────────

        lay.addWidget(card, 1)  # <-- IMPORTANT: stretch factor 1

        # -------- BUTTON ROW --------
        row = QHBoxLayout()

        back = QPushButton("Back")
        back.setProperty("nav", True)
        back.clicked.connect(lambda: self._goto_step(4))

        confirm = QPushButton("Confirm and Start WlanPoller")
        confirm.setProperty("nav", True)
        confirm.clicked.connect(self._start_run)

        

        row.addWidget(back)
        row.addStretch()
        
        row.addWidget(confirm)

        lay.addLayout(row)
        lay.addStretch()
        return w

    def _page_step7(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(12, 4, 12, 10)
        lay.setSpacing(4)

        run_header = QLabel("Run Log (CLI Output)")
        run_header.setStyleSheet("font-size:18px; font-weight:700; padding:2px 0;")
        lay.addWidget(run_header)

        self.run_card = QGroupBox()
        rlay = QVBoxLayout(self.run_card)
        rlay.setContentsMargins(8, 2, 8, 8)
        rlay.setSpacing(4)
        self.run_log = QTextEdit()
        self.run_log.setReadOnly(True)
        self.run_log.setMinimumHeight(120)
        self.run_log.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.run_log.setFont(QFont("Courier New", 13))
        rlay.addWidget(self.run_log)
        lay.addWidget(self.run_card, 3)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setTextVisible(True)
        self.progress.setAlignment(Qt.AlignCenter)
        self.progress.setFixedHeight(30)
        self.progress.setFont(QFont("Roboto", 11, QFont.Weight.Bold))
        self.progress.setFormat("%p%")
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #d1d5db;
                border-radius: 10px;
                background: #f3f6f9;
                color: #000000;
                text-align: center;
                padding: 2px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #34d399, stop:1 #16a34a);
                border-radius: 10px;
            }
        """)
        lay.addWidget(self.progress)

        # Bulk WLC mode only (>3 WLCs): "<done> / <total> WLCs completed"
        # note under the progress bar. Hidden for the manual (<=3 WLC) path.
        self.wlc_progress_label = QLabel("")
        self.wlc_progress_label.setAlignment(Qt.AlignCenter)
        self.wlc_progress_label.setStyleSheet("color:#374151; font-weight:600; font-size:11px; padding-top:2px;")
        self.wlc_progress_label.setVisible(False)
        lay.addWidget(self.wlc_progress_label)

        # ── AP TABLE ──────────────────────────────────────────
        self.ap_section = QWidget()
        ap_layout = QVBoxLayout(self.ap_section)
        ap_layout.setContentsMargins(0, 0, 0, 0)
        ap_layout.setSpacing(4)
        ap_layout.addWidget(QLabel("AP Table"))

        # 🔥 ALWAYS recreate correct table structure
        if self.operation_type == "AP Only":

            self.ap_table = QTableWidget()
            self.ap_table.setColumnCount(4)
            self.ap_table.setHorizontalHeaderLabels([
                "AP Name", "AP Model", "AP IP", "Status"
            ])

        else:

            self.ap_table = QTableWidget()
            self.ap_table.setColumnCount(5)
            self.ap_table.setHorizontalHeaderLabels([
                "WLC IP", "AP Name", "AP Model", "AP IP", "Status"
            ])
        header = self.ap_table.horizontalHeader()

        # ✅ Proper resize logic
        for i in range(self.ap_table.columnCount() - 1):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        header.setSectionResizeMode(self.ap_table.columnCount() - 1, QHeaderView.Stretch)
        header.setSectionResizeMode(self.ap_table.columnCount() - 1, QHeaderView.Stretch)
        if self.operation_type == "AP Only":
            self.ap_table.setColumnWidth(0, 220)
            self.ap_table.setColumnWidth(1, 160)
            self.ap_table.setColumnWidth(2, 160)
            self.ap_table.setColumnWidth(3, 500)
        else:
            self.ap_table.setColumnWidth(0, 160)  # WLC
            self.ap_table.setColumnWidth(1, 220)
            self.ap_table.setColumnWidth(2, 160)
            self.ap_table.setColumnWidth(3, 160)
            self.ap_table.setColumnWidth(4, 500)
        self.ap_table.verticalHeader().setVisible(False)
        self.ap_table.setWordWrap(True)
        self.ap_table.setTextElideMode(Qt.ElideNone)
        self.ap_table.verticalHeader().setDefaultSectionSize(34)
        self.ap_table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.ap_table.setAlternatingRowColors(True)
        self.ap_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.ap_table.setShowGrid(False)
        self.ap_table.setMinimumHeight(100)
        ap_layout.addWidget(self.ap_table)
        lay.addWidget(self.ap_section, 5)

        # ── VULNERABLE TABLE ──────────────────────────────────
        self.vuln_section = QWidget()
        vuln_layout = QVBoxLayout(self.vuln_section)
        vuln_layout.setContentsMargins(0, 0, 0, 0)
        vuln_layout.setSpacing(4)
        vuln_layout.addWidget(QLabel("Susceptible APs & Recovery Table"))

        self.vuln_table = QTableWidget(0, 6)
        self.vuln_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.vuln_table.setMinimumHeight(100)
        self.vuln_table.verticalHeader().setDefaultSectionSize(32)
        self.vuln_table.setHorizontalHeaderLabels(
            ["AP Name", "AP Model", "AP IP", "Active Boot Part", "Recovery", "Partition Note"]
        )
        header_v = self.vuln_table.horizontalHeader()
        header_v.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(4, QHeaderView.Stretch)
        header_v.setSectionResizeMode(5, QHeaderView.Stretch)
        self.vuln_table.setColumnWidth(0, 220)
        self.vuln_table.setColumnWidth(1, 160)
        self.vuln_table.setColumnWidth(2, 160)
        self.vuln_table.setColumnWidth(3, 130)
        self.vuln_table.setColumnWidth(4, 350)
        self.vuln_table.setColumnWidth(5, 350)
        self.vuln_table.setShowGrid(False)
        self.vuln_table.setAlternatingRowColors(True)
        # ── explicitly enable vertical scroll ────────────────
        self.vuln_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.vuln_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        vuln_layout.addWidget(self.vuln_table)
        lay.addWidget(self.vuln_section, 2)

        # ── RESULT SUMMARY ────────────────────────────────────
        self.results_summary = QTextEdit()
        self.results_summary.setReadOnly(True)
        self.results_summary.setMinimumHeight(50)
        lay.addWidget(QLabel("===== RESULT SUMMARY ====="))
        lay.addWidget(self.results_summary, 3)

        # ── ACTION BUTTONS ────────────────────────────────────
        actions = QHBoxLayout()
        self.btn_save_log = QPushButton("Save Run Log")
        self.btn_save_log.setProperty("nav", True)
        self.btn_save_log.clicked.connect(self._save_run_log)

        self.btn_export_vuln = QPushButton("Export Susceptible Table to Excel")
        self.btn_export_vuln.setProperty("nav", True)
        self.btn_export_vuln.clicked.connect(self._export_vuln_table)

        self.btn_view_logs = QPushButton("View Logs (Open Folder)", clicked=self._open_data_folder)
        self.btn_view_logs.setProperty("nav", True)
        self.chk_save_cronjob = QCheckBox("Save as CronJob")
        self.chk_save_cronjob.setStyleSheet("font-size:11px; color:#374151;")
        self.chk_save_cronjob.setToolTip(
            "Save this workflow to config.ini.\n"
            "Next time you launch the app it will auto-run this workflow."
        )
        # ✅ Trigger immediately on check — not deferred to _start_run
        self.chk_save_cronjob.stateChanged.connect(
            lambda state: self._save_cronjob() if state else None
        )

        self.chk_delete_cronjob = QCheckBox("Delete CronJob")
        self.chk_delete_cronjob.setStyleSheet("font-size:11px; color:#dc2626;")
        self.chk_delete_cronjob.setToolTip("Remove saved CronJob from config.ini")

        # Trigger immediately when checked (same pattern as save)
        self.chk_delete_cronjob.stateChanged.connect(
            lambda state: self._delete_cronjob() if state else None
        )
        self.chk_collect_again = QCheckBox("Collect Archive + WNCD Core again")
        self.chk_collect_again.setStyleSheet("font-size:11px; color:#0369a1;")
        self.chk_collect_again.setToolTip(
            "Resets the archive + WNCD core collection counter to 0.\n"
            "The next two workflow runs will collect archive + core files again."
        )
        self.chk_collect_again.stateChanged.connect(
            lambda state: self._reset_collect_counter() if state else None
        )
        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.close)
        self.btn_close.setProperty("nav", True)

        actions.addWidget(self.btn_save_log)
        actions.addWidget(self.btn_export_vuln)
        actions.addWidget(self.btn_view_logs)
        actions.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        actions.addWidget(self.chk_collect_again)
        actions.addWidget(self.chk_delete_cronjob)
        actions.addWidget(self.chk_save_cronjob)
        actions.addWidget(self.btn_close)
        lay.addLayout(actions)

        return w
    def _open_status_file(self):
        try:
            path = getattr(self, "last_status_file", "")

            if not path:
                QMessageBox.warning(self, "Missing", "Status summary log not available.")
                return

            folder = os.path.dirname(path)

            if not os.path.exists(folder):
                QMessageBox.warning(self, "Missing", "Log folder not found.")
                return

            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                os.system(f'open "{folder}"')
            else:
                os.system(f'xdg-open "{folder}"')

        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

    def _update_step7_visibility(self):
        """Show AP related widgets only if AP involved"""
        show_ap_related = self.operation_type in ("AP Only", "WLC & AP")

        if hasattr(self, "ap_cmd_section"):
            self.ap_section.setVisible(show_ap_related)

        if hasattr(self, "vuln_section"):
            self.vuln_section.setVisible(show_ap_related)

        if hasattr(self, "btn_export_vuln"):
            self.btn_export_vuln.setVisible(show_ap_related)

    def _page_parser(self) -> QWidget:
        w = QWidget();
        lay = QVBoxLayout(w)
        lay.addWidget(QLabel("<h2>Parser</h2>"))
        self.parser_mode = QComboBox();
        self.parser_mode.addItems(["WLC files", "AP files"])
        self.parser_pattern = QLineEdit();
        self.parser_pattern.setPlaceholderText("Enter regex or substring")
        btn = QPushButton("Search");
        btn.clicked.connect(self._run_parser);
        self.parser_out = QTextEdit();
        self.parser_out.setReadOnly(True)
        lay.addWidget(self.parser_mode);
        lay.addWidget(self.parser_pattern);
        lay.addWidget(btn);
        lay.addWidget(self.parser_out)
        row = QHBoxLayout()
        row.addStretch()

        self.btn_parser_close = QPushButton("Close")
        self.btn_parser_close.setProperty("nav", True)
        self.btn_parser_close.clicked.connect(self.close)

        row.addWidget(self.btn_parser_close)
        lay.addLayout(row)
        return w

    # ---------------- Actions / Helpers ----------------
    def _goto_step(self, idx: int):
        self.stack.setCurrentIndex(idx)

        # Force sidebar highlight
        self.sidebar.blockSignals(True)
        self.sidebar.setCurrentRow(idx)
        item = self.sidebar.item(idx)
        if item:
            item.setSelected(True)
        self.sidebar.blockSignals(False)

        try:
            self._refresh_visibility()
            self._update_step7_visibility()
        except Exception:
            pass

        if idx == 5:
            try:
                self._fill_preview()
            except Exception as e:
                if hasattr(self, "run_log"):
                    self.run_log.append(f"[DEBUG] preview build failed: {e}")
            try:
                self._check_iter_compatibility()
            except Exception:
                pass
    def _on_operation_change(self, value: str):
        self.operation_type = value

        # Clear AP list when switching away from AP Only
        if value != "AP Only" and not getattr(self, "headless_mode", False):
            self.ap_list_file = ""
            self.ap_list_path = ""
            if hasattr(self, "ap_path"):
                self.ap_path.setText("")

        # Show/hide Add+Remove WLC buttons
        if hasattr(self, "btn_add_wlc") and hasattr(self, "remove_btn_wlc"):
            if value in ("WLC Only", "WLC & AP"):
                self.btn_add_wlc.setVisible(True)
                self.remove_btn_wlc.setVisible(True)
            else:
                # AP Only — hide buttons and strip back to 1 WLC entry
                self.btn_add_wlc.setVisible(False)
                self.remove_btn_wlc.setVisible(False)
                if hasattr(self, "wlc_entries"):
                    while len(self.wlc_entries) > 1:
                        entry = self.wlc_entries.pop()
                        entry["widget"].deleteLater()

        # Show/hide WLC block
        if hasattr(self, "wlc_block"):
            self.wlc_block.setVisible(value != "AP Only")

        # Show/hide AP block
        if hasattr(self, "ap_block"):
            self.ap_block.setVisible(value != "WLC Only")

        # Downstream refresh
        self._refresh_step1()
        self._refresh_visibility()
        self._update_step7_visibility()
        self._update_workflow_dropdown()
    def _refresh_step1(self):

        is_ap_only = (self.operation_type == "AP Only")

        if hasattr(self, "ap_upload_row"):
            self.ap_upload_row.setVisible(is_ap_only)

        if hasattr(self, "ap_stats") and not is_ap_only:
            self.ap_stats.clear()

        # Always allow the button to be clicked.
        # Validation is handled in _step1_enter_credentials().
        if hasattr(self, "btn_step1_next"):
            self.btn_step1_next.setEnabled(True)
    def _browse_ap_list(self):

        path, _ = QFileDialog.getOpenFileName(
            self,
            "Upload AP List File (Format: AP Ip, AP Name)",
            "",
            "Text/CSV (*.txt *.csv);;All Files (*.*)"
        )

        if not path:
            return

        total_cnt = 0
        valid_cnt = 0
        invalid_cnt = 0
        duplicate_cnt = 0

        seen_ips = set()

        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            for i, line in enumerate(f, start=1):

                s = line.strip()
                if not s:
                    continue

                total_cnt += 1

                parts = [p.strip() for p in (s.split(",") if "," in s else s.split())]
                ip = parts[0]

                try:
                    socket.inet_pton(socket.AF_INET, ip)

                    if ip in seen_ips:
                        duplicate_cnt += 1
                    else:
                        valid_cnt += 1
                        seen_ips.add(ip)

                except Exception:
                    invalid_cnt += 1

        if valid_cnt == 0:
            QMessageBox.critical(self, "Invalid File", "No valid AP IP addresses found.")
            return

        os.makedirs(CONFD, exist_ok=True)

        self.ap_list_file = path
        self.ap_path.setText(path)
        self.ap_list_path = path

        # -------- COLORED STATUS TEXT --------
        stats_html = f"""
        <span style='color:#2563eb; font-weight:600;'>Total:</span> {total_cnt} |
        <span style='color:#16a34a; font-weight:600;'>Valid:</span> {valid_cnt} |
        <span style='color:#dc2626; font-weight:600;'>Invalid:</span> {invalid_cnt} |
        <span style='color:#f59e0b; font-weight:600;'>Duplicates:</span> {duplicate_cnt}
        """

        if hasattr(self, "ap_stats"):
            self.ap_stats.setText(stats_html)

        QMessageBox.information(
            self,
            "AP List Loaded",
            f"""File loaded successfully.

    Total APs: {total_cnt}
    Valid APs: {valid_cnt}
    Invalid APs: {invalid_cnt}
    Duplicate APs: {duplicate_cnt}
    """
        )

        self._refresh_step1()
    def _save_creds(self):
        if self.operation_type in ("WLC Only", "WLC & AP"):
            for i, entry in enumerate(getattr(self, "wlc_entries", [])):
                ip = entry["ip"].text().strip()
                if not ip:
                    continue
                if not is_ipv4_or_ipv6(ip):
                    QMessageBox.critical(self, "Invalid WLC IP", f"WLC {i+1}: invalid IP address.")
                    return

        if not self.ini:
            QMessageBox.warning(self, "Warning", "INI backend not available.")
            return

        if self.operation_type in ("WLC Only", "WLC & AP"):
            if self.ini and getattr(self, "wlc_entries", []):

                # CLEAR OLD WLC SECTIONS
                for sec in list(self.ini.cfg.sections()):
                    if sec.startswith("WLC"):
                        self.ini.cfg.remove_section(sec)

                # WRITE NEW WLC SECTIONS
                for i, e in enumerate(self.wlc_entries):

                    ip = e["ip"].text().strip()
                    if not ip:
                        continue   # ✅ skip empty WLC rows

                    section = "WLC" if i == 0 else f"WLC{i+1}"

                    self.ini.bulk_set(section, {
                        "wlc_ip": ip,
                        "wlc_user": e["user"].text().strip(),
                        "wlc_pasw": e["pasw"].text().strip()   # ✅ FIXED KEY
                    })

        # SAVE AP credentials
        if self.operation_type in ("WLC & AP", "AP Only"):
            try:
                self.ini.bulk_set("AP", {
                    "ap_user": self.ap_user.text().strip(),
                    "ap_pasw": self.ap_pass.text().strip(),
                    "ap_enable": self.ap_enable.text().strip()
                })
            except Exception:
                pass

        self.ini.save()

        print("DbgWpgui: Save Func Written to file : ", CONFD)
        print("DbgWpgui:Executable:", DATA_DIR)
        print("DbgWpgui:Base dir:", BASE_DIR)

        QMessageBox.information(self, "Saved", "Credentials saved to confd/config.ini")
    def _save_creds_silent(self):
        if not self.ini:
            return
        if self.operation_type in ("WLC Only", "WLC & AP"):
            try:
                if getattr(self, "wlc_entries", []):
                    # Only rewrite if at least WLC1 has an IP filled in
                    has_any_ip = any(
                        e["ip"].text().strip()
                        for e in self.wlc_entries
                    )
                    if not has_any_ip:
                        print("[HEADLESS] _save_creds_silent: no WLC IPs in widgets, skipping ini rewrite")
                    else:
                        # CLEAR OLD WLC SECTIONS (important to avoid stale entries)
                        # Only clear and rewrite if ALL entries have IPs (prevents corruption)
                        all_have_ip = all(
                            e["ip"].text().strip()
                            for e in self.wlc_entries
                            if e["ip"].text().strip()  # at least one non-empty
                        )
                        # CLEAR OLD WLC SECTIONS (important to avoid stale entries)
                        for sec in list(self.ini.cfg.sections()):
                            if sec.startswith("WLC"):
                                self.ini.cfg.remove_section(sec)
                        # WRITE EACH WLC AS ITS OWN SECTION
                        for i, e in enumerate(self.wlc_entries):
                            ip = e["ip"].text().strip()
                            if not ip:
                                continue  # skip empty entries
                            section = "WLC" if i == 0 else f"WLC{i+1}"
                            self.ini.bulk_set(section, {
                                "wlc_ip": ip,
                                "wlc_user": e["user"].text().strip(),
                                "wlc_pasw": e["pasw"].text()
                            })
            except Exception:
                pass
        if self.operation_type in ("WLC & AP", "AP Only"):
            try:
                self.ini.bulk_set("AP", {
                    "ap_user": self.ap_user.text().strip(),
                    "ap_pasw": self.ap_pass.text(),
                    "ap_enable": self.ap_enable.text()
                })
            except Exception:
                pass
        try:
            self.ini.save()
        except Exception:
            pass
    def _save_cronjob(self):
        """Save current run configuration into [CRONJOB] section of config.ini.
        Only operation_type and workflow are written to the ini file.
        All other runtime state is kept in memory only in self._cronjob_state.
        """
        if not self.ini:
            return
        try:
            wlc_cmds = getattr(self, "wlc_cmds", [])
            ap_cmds  = getattr(self, "ap_cmds",  [])

            if hasattr(self, "wlc_cmd_box"):
                box = [l.strip() for l in self.wlc_cmd_box.toPlainText().splitlines() if l.strip()]
                if box:
                    wlc_cmds = box
            if hasattr(self, "ap_cmd_box"):
                box = [l.strip() for l in self.ap_cmd_box.toPlainText().splitlines() if l.strip()]
                if box:
                    ap_cmds = box

            # Keep full runtime state in memory only
            self._cronjob_state = {
                "operation_type": self.operation_type,
                "workflow":        self.workflow,
                "ap_mode":         self.ap_mode,
                "ap_filter_mode":  self.ap_filter_mode,
                "site_tag":        self.site_tag,
                "model_group":     self.model_group,
                "ap_device":       self.ap_device,
                "ap_list_file":    getattr(self, "ap_list_file", ""),
                "wlc_cmds":        wlc_cmds,
                "ap_cmds":         ap_cmds,
            }
            try:
                for i, e in enumerate(getattr(self, "wlc_entries", [])):
                    ip = e["ip"].text().strip()
                    if ip:
                        sec = "WLC" if i == 0 else f"WLC{i+1}"
                        self._cronjob_state[f"cred_{sec}_ip"]   = ip
                        self._cronjob_state[f"cred_{sec}_user"] = e["user"].text().strip()
                        self._cronjob_state[f"cred_{sec}_pasw"] = e["pasw"].text().strip()
            except Exception:
                pass
            try:
                if hasattr(self, "ap_user"):
                    self._cronjob_state["cred_ap_user"]   = self.ap_user.text().strip()
                    self._cronjob_state["cred_ap_pasw"]   = self.ap_pass.text().strip()
                    self._cronjob_state["cred_ap_enable"] = self.ap_enable.text().strip()
            except Exception:
                pass

            # Write ONLY operation_type and workflow to the ini file
            if not self.ini.cfg.has_section("CRONJOB"):
                self.ini.cfg.add_section("CRONJOB")

            # Strip any verbose keys left over from previous saves
            for k in [
                "ap_mode", 
                "ap_device", "wlc_cmds", "ap_cmds",
                "cred_wlc_ip", "cred_wlc_user", "cred_wlc_pasw",
                "cred_wlc2_ip", "cred_wlc2_user", "cred_wlc2_pasw",
                "cred_wlc3_ip", "cred_wlc3_user", "cred_wlc3_pasw",
                "cred_ap_user", "cred_ap_pasw", "cred_ap_enable",
            ]:
                if self.ini.cfg.has_option("CRONJOB", k):
                    self.ini.cfg.remove_option("CRONJOB", k)

            self.ini.cfg.set("CRONJOB", "operation_type",          self.operation_type)
            self.ini.cfg.set("CRONJOB", "workflow",                self.workflow)
            self.ini.cfg.set("CRONJOB", "ap_filter_mode",          getattr(self, "ap_filter_mode", "NONE"))
            self.ini.cfg.set("CRONJOB", "site_tag",        getattr(self, "site_tag", ""))
            self.ini.cfg.set("CRONJOB", "model_group",     getattr(self, "model_group", "All AP Models"))
            self.ini.cfg.set("CRONJOB", "ap_list_file",    getattr(self, "ap_list_file", ""))
             # ── NEW: Iteration keys ──────────────────────────────
            self.ini.cfg.set("CRONJOB", "iterations_enabled",  str(getattr(self, "iterations_enabled",  False)).lower())
            self.ini.cfg.set("CRONJOB", "iteration_count",     str(getattr(self, "iteration_count",     1)))
            self.ini.cfg.set("CRONJOB", "iteration_interval",  str(getattr(self, "iteration_interval",  300)))
            
            # Initialize counter only if not already present (preserve existing count)
            if not self.ini.cfg.has_option("CRONJOB", "collect_archive_wnccore_count"):
                self.ini.cfg.set("CRONJOB", "collect_archive_wnccore_count", "0")
            if self.ini.cfg.has_section("ACTIVE_WORKFLOW"):
                self.ini.cfg.remove_section("ACTIVE_WORKFLOW")
            self.ini.save()
            # =========================
            # SAVE COMMANDS TO FILE (THIS IS THE MISSING LINK)
            # =========================
            try:
                os.makedirs(CONFD_DIR, exist_ok=True)

                wlc_file = os.path.join(CONFD_DIR, "wlc_cmds.txt")
                ap_file  = os.path.join(CONFD_DIR, "ap_cmds.txt")

                with open(wlc_file, "w", encoding="utf-8") as f:
                    f.write("\n".join(wlc_cmds))

                with open(ap_file, "w", encoding="utf-8") as f:
                    f.write("\n".join(ap_cmds))

                print("[CRONJOB] Commands saved to files")

            except Exception as e:
                print(f"[CRONJOB ERROR] Failed to save commands: {e}")
            if not getattr(self, "headless_mode", False):
                QMessageBox.information(
                    self, "CronJob Saved",
                    "Workflow saved to config.ini.\n\n"
                    "The next time you launch WlanPollerGUI it will automatically run this workflow."
                )
        except Exception as e:
            print(f"[CRONJOB] Save failed: {e}")

    def _delete_cronjob(self):
        """Remove [CRONJOB] section from config.ini."""
        if not self.ini:
            return
        if not self.ini.cfg.has_section("CRONJOB"):
            QMessageBox.information(self, "No CronJob", "No saved CronJob found in config.ini.")
            return
        reply = QMessageBox.question(
            self, "Delete CronJob",
            "Remove the saved CronJob?\n\nThe app will no longer auto-run on next launch.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.ini.cfg.remove_section("CRONJOB")
            if self.ini.cfg.has_section("ACTIVE_WORKFLOW"):
                self.ini.cfg.remove_section("ACTIVE_WORKFLOW")
            self.ini.save()
            QMessageBox.information(self, "Deleted", "CronJob removed from config.ini.")
    def _reset_collect_counter(self):
        """Reset collect_archive_wnccore_count to 0 in config.ini."""
        if not self.ini:
            return
        try:
            if not self.ini.cfg.has_section("CRONJOB"):
                self.ini.cfg.add_section("CRONJOB")
            self.ini.cfg.set("CRONJOB", "collect_archive_wnccore_count", "0")
            self.ini.save()
            QMessageBox.information(
                self,
                "Counter Reset",
                "Archive + WNCD Core collection counter reset to 0.\n\n"
                "The next two workflow runs will collect archive and core files again."
            )
            # Uncheck after confirming so user can click again later
            if hasattr(self, "chk_collect_again"):
                self.chk_collect_again.blockSignals(True)
                self.chk_collect_again.setChecked(False)
                self.chk_collect_again.blockSignals(False)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to reset counter: {e}")
    def _on_worker_log(self, text):
        """Append worker log and touch last_progress_time so watchdog knows worker is alive."""
        try:
            if hasattr(self, "run_log"):
                self.run_log.append(str(text))
            self.last_progress_time = time.time()
        except Exception:
            pass

    def _watchdog_check(self):
        """If the worker is running but no progress/log seen in 30 seconds, warn user."""
        try:
            if not hasattr(self, "worker") or self.worker is None:
                if getattr(self, "watchdog_timer", None):
                    self.watchdog_timer.stop()
                return

            # if worker not running anymore, stop watchdog
            try:
                if not getattr(self.worker, "isRunning", lambda: False)():
                    if getattr(self, "watchdog_timer", None):
                        self.watchdog_timer.stop()
                    return
            except Exception:
                pass

            last = getattr(self, "last_progress_time", None)
            timeout_sec = 120

            # Image download can be silent for a long time — give it more room
            ap_cmds = getattr(self, "ap_cmds", [])
            # If user typed commands, always use them regardless of mode dropdown
            if ap_cmds:
                self.ap_mode = "AP Custom Cmd List"
            is_image_run = any(
                "archive download-sw" in c.lower()
                or "sftp://" in c.lower()
                or "scp://" in c.lower()
                for c in ap_cmds
            )
            effective_timeout = 3600 if is_image_run else timeout_sec

            if last is None or (time.time() - last) > effective_timeout:
                if hasattr(self, "run_log"):
                    self.run_log.append(
                        f"[WATCHDOG] No progress or log for {effective_timeout}s. "
                        f"Worker may be waiting on a long-running device command."
                    )
                # optionally request interruption once
                try:
                    if hasattr(self.worker, "requestInterruption"):
                        self.worker.requestInterruption()
                except Exception:
                    pass
        except Exception:
            pass

    def _start_run(self):
        """
        Start the run without any confirmation dialog.
        Saves credentials silently, resets UI, builds the PollerWorker and starts it.
        """
        print('DbgWpgui: Inside Start_run..')
        # 🔥 FORCE CORRECT TABLE STRUCTURE BEFORE RUN

        if self.operation_type == "AP Only":

            self.ap_table.setColumnCount(4)
            self.ap_table.setHorizontalHeaderLabels([
                "AP Name", "AP Model", "AP IP", "Status"
            ])

        else:

            self.ap_table.setColumnCount(5)
            self.ap_table.setHorizontalHeaderLabels([
                "WLC IP", "AP Name", "AP Model", "AP IP", "Status"
            ])

        # 🔥 CLEAR OLD DATA
        self.ap_table.setRowCount(0)
        # Load shared WLC cmds (per-WLC overrides live in ini, read by engine)
        # Load shared WLC cmds (per-WLC overrides live in ini, read by engine)
        if hasattr(self, "wlc_cmd_box"):
            box_wlc = [l.strip() for l in self.wlc_cmd_box.toPlainText().splitlines() if l.strip()]
            if box_wlc:
                self.wlc_cmds = box_wlc
            # else: keep self.wlc_cmds as loaded from profile
        # Load AP cmds — but NEVER overwrite if workflow already built them
        # (Upload Files from AP and AP Flash Checker build cmds in _step3_proceed)
        _workflows_that_prebuild_cmds = ("Upload Files from AP", "AP Flash Checker", "TMP Cleanup + reload")
        if getattr(self, "workflow", "") not in _workflows_that_prebuild_cmds:
            if hasattr(self, "ap_cmd_box"):
                box_cmds = [l.strip() for l in self.ap_cmd_box.toPlainText().splitlines() if l.strip()]
                if box_cmds:
                    self.ap_cmds = box_cmds
        # Silent save
        try:
            self._save_creds_silent()
        except Exception:
            pass
        # ── Sync iteration state from Step6 widgets ──────────────
        try:
            self._sync_iter_state()
        except Exception:
            pass
         # CronJob is saved immediately on checkbox toggle (see _page_step7)
        self.ap_name_map = {}
        # Pre-populate name map so AP Only 2-column files (IP Name) display correctly
        if self.operation_type == "AP Only" and getattr(self, "ap_list_file", ""):
            try:
                with open(self.ap_list_file, "r", encoding="utf-8", errors="ignore") as _f:
                    for _line in _f:
                        _s = _line.strip()
                        if not _s:
                            continue
                        _parts = [p.strip() for p in (_s.split(",") if "," in _s else _s.split())]

                        _ip, _model, _name = normalize_ap_entry(_parts)

                        if _ip and _name:
                            self.ap_name_map[_ip] = _name
            except Exception:
                pass
        # Reset UI defensively
        for attr, op in (
                ("run_log", lambda w: w.clear()),
                ("ap_table", lambda w: w.setRowCount(0)),
                ("vuln_table", lambda w: w.setRowCount(0)),
                ("progress", lambda w: w.setValue(0)),
                ("results_summary", lambda w: w.clear()),
                ("wlc_progress_label", lambda w: (w.setText(""), w.setVisible(False))),
        ):

            if hasattr(self, attr):
                try:
                    op(getattr(self, attr))
                except Exception:
                    pass
        # Disconnect old worker signals before creating new one
        if hasattr(self, "worker") and self.worker is not None:
            try:
                self.worker.log.disconnect()
                self.worker.progress.disconnect()
                self.worker.ap_update.disconnect()
                self.worker.finished_ok.disconnect()
                self.worker.failed.disconnect()
                self.worker.wlc_progress.disconnect()
            except Exception:
                pass
            self.worker = None
        # Decide ap_list_file and ap_cmds
        ap_list_file = getattr(self, "ap_list_file", "")


        if self.operation_type != "AP Only":
            ap_list_file = ""

        if getattr(self, "workflow", "") == "AP Flash Checker":
            ap_cmds = getattr(self, "ap_cmds", []) or [
                "show clock",
                "show version",
                "show flash",
                "show flash | i cnssdaemon.log",
                "show boot",
                "show filesystems",
                "show image integrity",
            ]
        else:
            ap_cmds = getattr(self, "ap_cmds", [])



        # Build the worker
        try:
            self.worker = PollerWorker(
                operation_type=self.operation_type,
                workflow=self.workflow,
                wlc_cmds=getattr(self, "wlc_cmds", []),
                ap_cmds=ap_cmds,
                ap_filter_mode=getattr(self, "ap_filter_mode", "NONE"),
                site_tag=getattr(self, "site_tag", ""),
                model_group=getattr(self, "model_group", "All AP Models"),
                ap_device=getattr(self, "ap_device", "cos_qca"),
                ap_list_file=ap_list_file,
                ap_mode=getattr(self, "ap_mode", "AP Custom Cmd List"),
                iterations_enabled  = getattr(self, "iterations_enabled",  False),
                iteration_count     = getattr(self, "iteration_count",     1),
                iteration_interval  = getattr(self, "iteration_interval",  300),
                )
            
            self.worker.enable_tmp_cleanup = getattr(self, "enable_tmp_cleanup", False)
            self.worker.enable_reload = getattr(self, "enable_reload", False)
            self.worker.enable_debug_collection = getattr(self, "enable_debug_collection", False)
            self.worker.test_after_iteration = getattr(self, "test_after_iteration", False)

            # Bulk WLC mode (>3 WLCs) — only active when the Bulk Upload radio
            # is selected and a WLC list has been parsed from a file. Manual
            # (<=3 WLC) runs always get an empty list here, so PollerWorker.run()
            # falls back to its existing config.ini-based path unchanged.
            if getattr(self, "wlc_mode_bulk_radio", None) and self.wlc_mode_bulk_radio.isChecked():
                self.worker.wlc_bulk_list = getattr(self, "wlc_bulk_list", [])
                self.worker.bulk_wlc_user = (
                    self.wlc_entries[0]["user"].text().strip() if getattr(self, "wlc_entries", []) else ""
                )
                self.worker.bulk_wlc_pasw = (
                    self.wlc_entries[0]["pasw"].text() if getattr(self, "wlc_entries", []) else ""
                )
            else:
                self.worker.wlc_bulk_list = []
        except Exception as e:
            QMessageBox.critical(self, "Worker Error", f"Failed to create worker: {e}")
            return

        # Prepare watchdog state
        self.last_progress_time = time.time()
        if getattr(self, "watchdog_timer", None) is None:
            self.watchdog_timer = QTimer(self)
            self.watchdog_timer.setInterval(10000)  # check every 10 seconds
            self.watchdog_timer.timeout.connect(self._watchdog_check)

        # Hook signals with small wrappers that update last_progress_time
        try:
            # log -> wrapper that updates last_progress_time

            self.worker.log.connect(self._on_worker_log)
        except Exception:
            pass

        try:
            # progress -> update bar and last_progress_time

            def _progress_cb(pct):
                try:
                    if hasattr(self, "progress"):
                        self.progress.setValue(pct)
                    self.last_progress_time = time.time()
                except Exception:
                    pass

            self.worker.progress.connect(self._ui_progress_update)
        except Exception:
            pass

        try:

            self.worker.ap_update.connect(self._on_ap_update)
        except Exception:
            pass

        try:
            self.worker.wlc_progress.connect(self._on_wlc_progress)
        except Exception:
            pass

        # finished: stop watchdog and delegate to your existing _on_finished
        try:
            if hasattr(self.worker, "finished_ok") and callable(getattr(self.worker.finished_ok, "connect", None)):
                def _on_finished_wrapper(summary):
                    try:
                        if getattr(self, "watchdog_timer", None):
                            self.watchdog_timer.stop()
                    except Exception:
                        pass
                    self.run_in_progress = False
                    if hasattr(self, "btn_close"):
                        self.btn_close.setEnabled(True)
                    try:
                        self._on_finished(summary)
                    except Exception:
                        pass

                    # IMPORTANT: re-enable close button (missing piece)
                    try:
                        if hasattr(self, "btn_close"):
                            self.btn_close.setEnabled(True)
                    except Exception:
                        pass

                self.worker.finished_ok.connect(_on_finished_wrapper)
        except Exception:
            pass

        # failed: stop watchdog and show error (also re-enable sidebar)
        try:
            if hasattr(self.worker, "failed") and callable(getattr(self.worker.failed, "connect", None)):
                def _on_fail(e):
                    self.run_in_progress = False

                    try:
                        if getattr(self, "watchdog_timer", None):
                            self.watchdog_timer.stop()
                    except Exception:
                        pass
                    try:
                        if hasattr(self, "run_log"):
                            self.run_log.append("[WORKER FAILED] " + str(e))
                    except Exception:
                        pass
                    if not getattr(self, "headless_mode", False):
                        try:
                            QMessageBox.critical(self, "Run Failed", e)
                        except Exception:
                            pass
                    else:
                        print(f"[HEADLESS] Run failed: {e}")
                    try:
                        if hasattr(self, "sidebar"):
                            self.sidebar.setEnabled(True)
                    except Exception:
                        pass

                    if hasattr(self, "btn_close"):
                        self.btn_close.setEnabled(True)
                    self.run_in_progress = False
                    
                self.worker.failed.connect(_on_fail)
        except Exception:
            pass

        # When thread starts -> append a visible message
        try:
            self.worker.started.connect(lambda: self.run_log.append("[WORKER] threads started"))
        except Exception:
            pass

        # Optionally lock navigation / sidebar while run is active
        try:
            if hasattr(self, "sidebar"):
                self.sidebar.setEnabled(False)
        except Exception:
            pass

        # Start watchdog and the worker
        try:
            self.last_progress_time = time.time()
            self.watchdog_timer.start()
            self._goto_step(6)
            self._inject_run_preview_into_log()
        except Exception:
            try:
                self.stack.setCurrentIndex(6)
            except Exception:
                pass
        # disable close while execution running
        if hasattr(self, "btn_close"):
            self.btn_close.setEnabled(False)
        self.run_in_progress = True
        if self.operation_type == "AP Only" and getattr(self, "ap_list_file", ""):
            try:
                with open(self.ap_list_file, "r", encoding="utf-8", errors="ignore") as f:
                    lines = [l for l in f if l.strip()]
            except Exception as e:
                print(f"[HEADLESS] Warning: could not pre-read AP list: {e}")
            # Copy AP list file to ap_ip_list_all.txt for AP Only mode
            try:
                import shutil
                dest = os.path.join(CONFD, "ap_ip_list_all.txt")
                # Only copy if dest doesn't exist or is empty
                
                shutil.copy2(self.ap_list_file, dest)
            except Exception:
                pass
        if self.operation_type != "AP Only":
            try:
                open(os.path.join(CONFD, "ap_ip_list_all.txt"), "w").close()
            except Exception:
                pass
        try:
            start_time = datetime.now().strftime("%H:%M:%S")

            msg = (
                "===== STARTING WLAN POLLER =====\n"
                f"Operation: {self.operation_type}\n"
                f"Workflow: {self.workflow}\n"
                f"Start Time: {start_time}\n"
            )
            self.run_log.append(msg)
            self.worker.start()
            if hasattr(self, "run_log"):
                self.run_log.append("[WORKER] start() called")
        except Exception as e:
            QMessageBox.critical(self, "Start Error", f"Failed to start worker: {e}")
            if hasattr(self, "sidebar"):
                self.sidebar.setEnabled(True)

    def _step2_proceed(self):
        if self.operation_type!= "AP Only":
            missing = self._validate_all_wlcs()

            if missing:
                QMessageBox.warning(
                    self,
                    "Missing WLC Credentials",
                    f"Please fill credentials for:\n{', '.join(missing)}"
                )
                return  # 🚫 BLOCK navigation
        if self.operation_type in ("WLC & AP", "AP Only"):
            if not self.ap_user.text().strip() or not self.ap_pass.text().strip():
                QMessageBox.warning(self, "Missing AP Credentials", "Please fill AP Username and Password.")
                return
        self._update_workflow_dropdown()
        
        # ✅ Only move if validation passed
        self._goto_step(2)
    def _step3_proceed(self):
        """
        Proceed from Step3 (Workflow).
        - AP Flash Checker: auto-fill AP commands and SKIP Step4 -> Step5 (Filters).
        - AP Image Download: set ap_mode and SKIP Step4 -> Step5 (user must provide FTP elsewhere).
        - Custom CLI Commands: go to Step4 so user can edit WLC/AP cmd lists.
        """
        
        # Prefer the widget value if present, else use stored state
        if hasattr(self, "workflow_dd") and callable(getattr(self.workflow_dd, "currentText", None)):
            wf = self.workflow_dd.currentText()
        else:
            wf = getattr(self, "workflow", None)

        # Persist selection in UI state
        self.workflow = wf
        # ==============================
        # AP CLEANUP + RELOAD (NEW)
        # ==============================
        if wf == "TMP Cleanup + reload":

            self.ap_cmds = ["%reload%"]
            self.ap_mode = "AP Custom Cmd List"
            self.ap_filter_mode = "NONE"

            self._fill_preview()

            self.sidebar.blockSignals(True)
            self.sidebar.setCurrentRow(5)
            self.sidebar.blockSignals(False)
            self.stack.setCurrentIndex(5)

            return

            
        # AP Flash Checker -> set default AP cmds and skip Step4
        # AP Flash Checker -> set default AP cmds and skip Step4
        if wf == "AP Flash Checker":

            self.ap_cmds = [
                "show clock",
                "show version",
                "show flash",
                "show flash | i cnssdaemon.log",
                "show boot",
                "show filesystem",
                "show image integrity"
                
            ]
            

            self.ap_mode = getattr(self, "ap_mode", "AP Custom Cmd List")

            # 🔥 NEW: STORE USER OPTIONS
            self.enable_tmp_cleanup = False
            self.enable_reload = False
                
            
            # ---- FLOW CONTROL ----
            if self.operation_type == "AP Only":

                self.ap_filter_mode = "NONE"

                self._fill_preview()
                self._goto_step(5)   # Step6 Preview

            else:
                self._goto_step(4)

            return
        # AP Image Download -> remember mode and skip Step4
        
# AP Image Download -> remember mode and skip Step4
        if wf == "AP Image Download":
            self.ap_mode = "AP Image Download"
            self._goto_step(4)
            return

        # Upload Files from AP -> validate, build commands, skip Step4
        if wf == "Upload Files from AP":
            server_ip = self.upload_server_ip_field.text().strip() if hasattr(self, "upload_server_ip_field") else ""

            if not server_ip:
                QMessageBox.critical(self, "Missing", "Enter the Server IP address for file upload.")
                return

            if not is_ipv4_or_ipv6(server_ip):
                QMessageBox.critical(self, "Invalid IP", "Server IP is not a valid IPv4/IPv6 address.")
                return

            proto     = self.upload_proto_dd.currentText()     if hasattr(self, "upload_proto_dd")     else "TFTP"
            file_type = self.upload_file_type_dd.currentText() if hasattr(self, "upload_file_type_dd") else "Syslogs"

            if proto == "SFTP":
                sftp_user = getattr(self, "upload_sftp_user", None)
                sftp_pass = getattr(self, "upload_sftp_pass", None)
                if not (sftp_user and sftp_user.text().strip()):
                    QMessageBox.critical(self, "Missing", "Enter SFTP username.")
                    return
                if not (sftp_pass and sftp_pass.text()):
                    QMessageBox.critical(self, "Missing", "Enter SFTP password.")
                    return

            # Auto-build the AP command — user never touches Step4 for this workflow
            self.ap_cmds  = self._build_upload_cmds(file_type, proto, server_ip)
            self.ap_mode  = "AP Custom Cmd List"

            # Persist state for preview
            self.upload_file_type   = file_type
            self.upload_server_ip_val = server_ip
            self.upload_proto       = proto

            if self.operation_type == "AP Only":
                # No filters needed — go straight to Preview
                self.ap_filter_mode = "NONE"
                self._fill_preview()
                self._goto_step(5)
            else:
                # WLC & AP — still allow site/model filter
                self._goto_step(4)
            return
        if wf == "AP Datapath Queue Mon":
            self.ap_cmds = []
            self.ap_filter_mode = "NONE"
            self._fill_preview()
            self._goto_step(5)   # jump straight to Step6 Preview — no Step4, no Step5
            return
        # Default: Custom CLI Commands -> show Step4
        if self.workflow == "Client Stuck In Auth Loop":
            self._goto_step(5)   # skip Step4 → go to Step5
        else:
            self._goto_step(3)   # normal flow
        

    def _step4_proceed(self):
        """
        Step4 → decide next navigation step.
        """
        if self.workflow == "AP Datapath Queue Mon":
            self.ap_cmds = []
            self._fill_preview()
            self._goto_step(5)
            return
        # Read WLC commands
        try:
            if hasattr(self, "wlc_cmd_box"):
                self.wlc_cmds = [
                    l.strip() for l in self.wlc_cmd_box.toPlainText().splitlines() if l.strip()
                ]
        except Exception:
            self.wlc_cmds = getattr(self, "wlc_cmds", [])

        # Read AP commands
        try:
            if hasattr(self, "ap_cmd_box"):
                self.ap_cmds = [
                    l.strip() for l in self.ap_cmd_box.toPlainText().splitlines() if l.strip()
                ]
        except Exception:
            self.ap_cmds = getattr(self, "ap_cmds", [])

        # ---------------- WLC ONLY ----------------
        if self.operation_type == "WLC Only":

            if not self.wlc_cmds:
                QMessageBox.critical(self, "Missing", "Enter WLC Cmd List.")
                return

            self._fill_preview()
            self._goto_step(5)
            return

        # ---------------- AP ONLY ----------------
        if self.operation_type == "AP Only":

            # AP commands required only for Custom Cmd mode
            if getattr(self, "ap_mode", "") == "AP Custom Cmd List":
                if not self.ap_cmds:
                    QMessageBox.critical(self, "Missing", "Enter AP Cmd List.")
                    return

            if getattr(self, "ap_mode", "") == "AP Image Download":

                confirm = QMessageBox.warning(
                    self,
                    "Verify Image Before Proceeding",
                    "⚠️  Please double-check before continuing:\n\n"
                    "  • AP Model in your AP list file\n"
                    "  • Image filename matches that model\n\n"
                    "Wrong image will cause transfer failure.\n\n"
                    "Proceed?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if confirm != QMessageBox.Yes:
                    return

            self._fill_preview()
            self._goto_step(5)
            return

        # ---------------- WLC & AP ----------------
        if self.operation_type == "WLC & AP":

            if not self.wlc_cmds:
                QMessageBox.critical(self, "Missing", "Enter WLC Cmd List.")
                return

            if not self.ap_cmds:
                QMessageBox.critical(self, "Missing", "Enter AP Cmd List.")
                return

            if getattr(self, "ap_mode", "") == "AP Image Download":

                confirm = QMessageBox.warning(
                    self,
                    "Verify Image Before Proceeding",
                    "⚠️  Please double-check before continuing:\n\n"
                    "  • AP Model from WLC AP summary\n"
                    "  • Image filename matches that model\n\n"
                    "Wrong image will cause transfer failure.\n\n"
                    "Proceed?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if confirm != QMessageBox.Yes:
                    return

            # go to Step5 Filters
            self._goto_step(4)

    def _step4_save(self):

        

            # ---------------- READ WLC COMMANDS ----------------
            wlc_cmds = []
            if hasattr(self, "wlc_cmd_box"):
                try:
                    wlc_cmds = [l.strip() for l in self.wlc_cmd_box.toPlainText().splitlines() if l.strip()]
                except Exception:
                    wlc_cmds = []

            # ---------------- READ AP COMMANDS ----------------
            ap_cmds = []
            if hasattr(self, "ap_cmd_box"):
                try:
                    ap_cmds = [l.strip() for l in self.ap_cmd_box.toPlainText().splitlines() if l.strip()]
                except Exception:
                    ap_cmds = []

# Per-WLC boxes are optional — empty means use the shared list above

            # ---------------- FTP VALIDATION (KEEP YOUR LOGIC) ----------------
            ap_mode = None
            if hasattr(self, "ap_mode_dd") and callable(getattr(self.ap_mode_dd, "currentText", None)):
                ap_mode = self.ap_mode_dd.currentText()
            else:
                ap_mode = getattr(self, "ap_mode", "AP Custom Cmd List")

            if ap_mode == "AP Image Download":
                ftp_missing = False

                ftp_user = getattr(self, "ftp_user", None)
                ftp_pasw = getattr(self, "ftp_pasw", None)
                ftp_addr = getattr(self, "ftp_addr", None)
                ftp_path = getattr(self, "ftp_path", None)
                scp_port = getattr(self, "scp_port", None)

                if not (ftp_user and ftp_user.text().strip()):
                    ftp_missing = True
                if not (ftp_pasw and ftp_pasw.text()):
                    ftp_missing = True

                if ftp_missing:
                    QMessageBox.critical(self, "FTP Missing", "FTP(SFTP) fields are mandatory for AP Image Download.")
                    return False

                if self.ini:
                    try:
                        self.ini.bulk_set("FTP", {
                            "ftp_addr": ftp_addr.text().strip() if ftp_addr else "",
                            "ftp_path": ftp_path.text().strip() if ftp_path else "",
                            "ftp_user(sftp)": ftp_user.text().strip(),
                            "ftp_pasw(sftp)": ftp_pasw.text(),
                            "scp_port": scp_port.text().strip() if scp_port and scp_port.text().strip() else "22"
                        })
                        self.ini.save()
                    except Exception:
                        QMessageBox.warning(self, "Warning", "Failed to save FTP settings.")

            # ---------------- CREATE CONFIG DIR ----------------
            try:
                os.makedirs(CONFD, exist_ok=True)
            except Exception:
                QMessageBox.critical(self, "Disk Error", f"Unable to create config folder: {CONFD}")
                return False

            # ---------------- SAVE GLOBAL WLC FILE (OPTIONAL BACKUP) ----------------
            if wlc_cmds:
                try:
                    with open(os.path.join(CONFD, "cmdlist_wlc.txt"), "w", encoding="utf-8") as f:
                        f.write("\n".join(wlc_cmds) + "\n")
                except Exception as e:
                    QMessageBox.warning(self, "Save Error", f"Failed to write WLC cmd list: {e}")

            # ---------------- SAVE AP COMMAND FILE ----------------
            if ap_cmds:
                fname = "cmdlist_cos_qca.txt"
                try:
                    dev = getattr(self, "ap_device", "cos_qca")
                    if dev == "cos":
                        fname = "cmdlist_cos.txt"
                    elif dev == "cos_bcm":
                        fname = "cmdlist_cos_bcm.txt"
                except Exception:
                    pass

                try:
                    with open(os.path.join(CONFD, fname), "w", encoding="utf-8") as f:
                        f.write("\n".join(ap_cmds) + "\n")
                except Exception as e:
                    QMessageBox.warning(self, "Save Error", f"Failed to write AP cmd list: {e}")

            # ---------------- UPDATE INTERNAL STATE ----------------
            self.wlc_cmds = wlc_cmds
            self.ap_cmds = ap_cmds

            # ---------------- SAVE PER-WLC COMMANDS (FINAL SOURCE) ----------------
            if hasattr(self, "per_wlc_cmd_boxes") and self.ini:

                # 🔥 CLEAR OLD SECTIONS FIRST
                for sec in list(self.ini.cfg.sections()):
                    if sec.endswith("_CMDS"):
                        self.ini.cfg.remove_section(sec)

                # 🔥 SAVE ALL (STRICT)
                # Save per-WLC overrides; blank box = remove section (use shared list)
                for section, box in self.per_wlc_cmd_boxes.items():
                    cmds = box.toPlainText().strip()
                    cmds_section = f"{section}_CMDS"
                    if cmds:
                        if not self.ini.cfg.has_section(cmds_section):
                            self.ini.cfg.add_section(cmds_section)
                        self.ini.cfg.set(cmds_section, "cmds", cmds)
                    else:
                        if self.ini.cfg.has_section(cmds_section):
                            self.ini.cfg.remove_section(cmds_section)

                try:
                    self.ini.save()
                except Exception:
                    QMessageBox.warning(self, "Warning", "Failed to save WLC command sections.")

            # ---------------- SAVE PROTOCOL INFO ----------------
            if hasattr(self, "proto_dd") and self.ini:
                try:
                    self.ini.bulk_set("FTP", {
                        "ftp_proto": self.proto_dd.currentText(),
                        "ftp_user": self.ftp_user.text().strip() if hasattr(self, "ftp_user") else "",
                        "ftp_pasw": self.ftp_pasw.text() if hasattr(self, "ftp_pasw") else "",
                    })
                    self.ini.save()
                except Exception:
                    pass

            QMessageBox.information(self, "Saved", "Cmd lists (and FTP(SFTP) details if provided) saved under confd/")

            return True
    def _save_run_log(self):
        """
        Save the Run Log contents to data/WlanPoller_RunLog_YYYYMMDD_HHMMSS.txt.
        Defensive: works even if run_log widget is not present.
        """
        try:
            folder = DATA_DIR
            os.makedirs(folder, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fn = os.path.join(folder, f"WlanPoller_RunLog_{ts}.txt")
            txt = ""
            if hasattr(self, "run_log") and self.run_log is not None:
                try:
                    txt = self.run_log.toPlainText()
                except Exception:
                    # fallback: read whatever attribute might exist
                    try:
                        txt = str(self.run_log)
                    except Exception:
                        txt = ""
            with open(fn, "w", encoding="utf-8") as f:
                f.write(txt)
            QMessageBox.information(self, "Saved", f"Run log saved to: {fn}")
        except Exception as e:
            # Show a friendly error if saving fails
            try:
                QMessageBox.warning(self, "Save Failed", f"Unable to save run log: {e}")
            except Exception:
                # last resort: print to console
                print("Unable to save run log:", e)
    def _save_run_log_headless(self):
        """Silently save run log to data/ folder — used when cronjob auto-run completes."""
        try:
            folder = DATA_DIR
            os.makedirs(folder, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fn = os.path.join(folder, f"WlanPoller_CronJobLog_{ts}.txt")
            txt = ""
            if hasattr(self, "run_log") and self.run_log is not None:
                try:
                    txt = self.run_log.toPlainText()
                except Exception:
                    txt = ""
            with open(fn, "w", encoding="utf-8") as f:
                f.write(txt)
            print(f"[CRONJOB] Log saved to: {fn}")
        except Exception as e:
            print(f"[CRONJOB] Log save failed: {e}")
    def _on_ap_mode_changed(self, text: str):
        self.ap_mode = text
        is_image = (text == "AP Image Download")

        # cmd box always visible when AP is involved
        if hasattr(self, "ap_cmd_box"):
            self.ap_cmd_box.setVisible(True)

        # ftp_group only visible when Image Download selected
        if hasattr(self, "ftp_group"):
            self.ftp_group.setVisible(is_image)

        self._refresh_visibility()
    def _export_ap_table(self):
        """
        Export the AP Table to an Excel file under data/.
        Defensive: checks for openpyxl and widget presence.
        """
        if Workbook is None:
            QMessageBox.warning(self, "Missing", "openpyxl is not installed. Run: pip install openpyxl")
            return

        # Quick helper: safely read cell text
        def _cell_text(table, r, c):
            try:
                it = table.item(r, c)
                return it.text() if (it and it.text()) else ""
            except Exception:
                return ""

        # Build rows from AP table (single pass)
        rows = []
        if hasattr(self, "ap_table"):
            try:
                for r in range(self.ap_table.rowCount()):
                    # Expect columns: 0=AP Name, 1=AP IP, 2=AP Model, 3=Status
                    # If the table doesn't have 4 columns yet, we still read safely
                    # when building rows for export (single pass)
                    name = _cell_text(self.ap_table, r, 0)
                    model = _cell_text(self.ap_table, r, 1)
                    ip = _cell_text(self.ap_table, r, 2)
                    status = _cell_text(self.ap_table, r, 3)
                    rows.append([name, model, ip, status])
                    # and set headers accordingly: ["AP Name","AP Model","AP IP","Status"]

            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed reading AP table: {e}")
                return
        else:
            QMessageBox.information(self, "No Data", "AP Table not found / empty.")
            return

        # Prepare Excel file
        folder = DATA_DIR
        os.makedirs(folder, exist_ok=True)
        fn = os.path.join(folder, f"WlanPoller_AP_Table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "AP Results"

            headers = ["AP Name", "AP IP", "AP Model", "Status"]
            ws.append(headers)

            # header styling
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = XLFont(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # write rows (one append per row is fine)
            for rdata in rows:
                ws.append(rdata)

            # freeze & autofilter
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

            # nice column widths
            ws.column_dimensions["A"].width = 32  # AP Name
            ws.column_dimensions["B"].width = 18  # AP IP
            ws.column_dimensions["C"].width = 18  # AP Model
            ws.column_dimensions["D"].width = 80  # Status

            wb.save(fn)
            QMessageBox.information(self, "Excel Exported", f"Saved: {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file: {e}")

    def _export_vuln_table(self):
        """
        Export the Vulnerable APs table to an Excel file under data/.
        Defensive: checks for openpyxl and widget presence.
        """
        if Workbook is None:
            QMessageBox.warning(self, "Missing", "openpyxl is not installed. Run: pip install openpyxl")
            return

        # Build rows from Vulnerable table
        rows = []
        if hasattr(self, "vuln_table"):
            try:
                for r in range(self.vuln_table.rowCount()):
                    name = self.vuln_table.item(r, 0).text() if self.vuln_table.item(r, 0) else ""
                    model = self.vuln_table.item(r, 1).text() if self.vuln_table.item(r, 1) else ""
                    ip = self.vuln_table.item(r, 2).text() if self.vuln_table.item(r, 2) else ""
                    boot_part = self.vuln_table.item(r, 3).text() if self.vuln_table.item(r, 3) else ""
                    recovery = self.vuln_table.item(r, 4).text() if self.vuln_table.item(r, 4) else ""
                    part_note = self.vuln_table.item(r, 5).text() if self.vuln_table.item(r, 5) else ""
                    rows.append([name, model, ip, boot_part, recovery, part_note])
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed reading Vulnerable table: {e}")
                return
        else:
            QMessageBox.information(self, "No Data", "Vulnerable APs table not found / empty.")
            return

        # Prepare Excel file
        folder = DATA_DIR
        os.makedirs(folder, exist_ok=True)
        fn = os.path.join(folder, f"WlanPoller_Susceptible_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Susceptible APs"

            headers = ["AP Name", "AP Model", "AP IP", "Active Boot Part", "Recovery", "Partition Note"]
            ws.append(headers)

            # header styling
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = XLFont(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # write rows
            for rdata in rows:
                ws.append(rdata)

            # freeze & autofilter
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

            # nice column widths
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 60
            ws.column_dimensions["F"].width = 60

            wb.save(fn)
            QMessageBox.information(self, "Excel Exported", f"Saved: {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file: {e}")

    def _open_data_folder(self):

        folder = str(DATA_DIR)


        # Prefer last run folder if available
        if hasattr(self, "last_status_file"):
            try:
                folder = os.path.dirname(self.last_status_file)
            except Exception:
                pass

        try:
            os.makedirs(folder, exist_ok=True)

            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                os.system(f'open "{folder}"')
            else:
                os.system(f'xdg-open "{folder}"')

        except Exception as e:
            QMessageBox.warning(self, "Open Folder Failed", str(e))
    def _on_workflow_change(self, v: str):
        self.workflow = v
        is_datapath_mon = (v == "AP Datapath Queue Mon")
        if hasattr(self, "datapath_mon_widget"):
            self.datapath_mon_widget.setVisible(is_datapath_mon)
        if hasattr(self, "ap_cmd_box"):
            self.ap_cmd_box.setVisible(not is_datapath_mon)
        if hasattr(self, "ap_mode_dd"):
            self.ap_mode_dd.setEnabled(not is_datapath_mon)
        if hasattr(self, "ftp_group") and is_datapath_mon:
            self.ftp_group.setVisible(False)
        # Upload config visibility
        if hasattr(self, "upload_config_widget"):
            self.upload_config_widget.setVisible(v == "Upload Files from AP")

        
        

        # 🔥 FLASH CHECKER VISIBILITY
        if hasattr(self, "flash_options_widget"):

            if self.operation_type == "AP Only" and v == "AP Flash Checker":
                self.flash_options_widget.setVisible(True)

                if hasattr(self, "tmp_hint_label"):
                    self.tmp_hint_label.setText(
                        "Note: /tmp cleanup will run only on APs with >60% usage."
                    )

            else:
                self.flash_options_widget.setVisible(False)
    # ... (remaining methods: _step3_proceed, _on_ap_mode_changed, _step4_save, _step4_proceed, _enforce_one_filter,
    # _update_ap_device_from_model, _step5_preview are defined earlier in file - kept unchanged for brevity) ...
    # For completeness they are implemented above in the full content.
        self._check_iter_compatibility()
    def _on_upload_proto_changed(self, proto: str):
        """Show/hide SFTP credential fields in the upload config block."""
        is_sftp = (proto == "SFTP")
        for attr in ("upload_sftp_user", "upload_sftp_pass",
                     "upload_sftp_user_label", "upload_sftp_pass_label"):
            if hasattr(self, attr):
                getattr(self, attr).setVisible(is_sftp)
    def _on_upload_proto_changed(self, proto: str):
        """Show/hide SFTP credential fields in the upload config block."""
        is_sftp = (proto == "SFTP")
        for attr in ("upload_sftp_user", "upload_sftp_pass",
                     "upload_sftp_user_label", "upload_sftp_pass_label"):
            if hasattr(self, attr):
                getattr(self, attr).setVisible(is_sftp)
    def _build_upload_cmds(self, file_type: str, proto: str, server_ip: str) -> List[str]:
        """
        Build the AP CLI command(s) for the Upload Files workflow.

        TFTP example: copy syslogs tftp: 192.168.0.10/
        SFTP example: ip sftp username admin
                      ip sftp password <pass>
                      copy syslogs sftp: 192.168.0.10/
        """
        type_map = {
            "Syslogs":       "syslogs",
            "Core Files":    "core:",
            "CrashFiles":    "crashinfo:",
            "SupportBundle": "support-bundle",
        }
        source = type_map.get(file_type, "syslogs")

        cmds = []

        if proto == "SFTP":
            user = self.upload_sftp_user.text().strip() if hasattr(self, "upload_sftp_user") else ""
            pasw = self.upload_sftp_pass.text() if hasattr(self, "upload_sftp_pass") else ""
            if user:
                cmds.append(f"ip sftp username {user}")
            if pasw:
                cmds.append(f"ip sftp password {pasw}")
            cmds.append(f"copy {source} sftp://{server_ip}/")
        else:
            cmds.append(f"copy {source} tftp://{server_ip}/")

        return cmds
    def _on_proto_changed(self, proto):
        is_sftp = (proto == "SFTP")

        # Server IP and Remote Path — only needed for SFTP
        for attr in ("ftp_addr", "ftp_path", "ftp_user", "ftp_pasw",
                     "ftp_user_label", "ftp_pasw_label"):
            if hasattr(self, attr):
                getattr(self, attr).setVisible(is_sftp)

        # Also hide/show the form labels via the ftp_layout rows
        if hasattr(self, "ftp_group") and self.ftp_group.layout():
            lay = self.ftp_group.layout()
            for i in range(lay.rowCount()):
                label_item = lay.itemAt(i, QFormLayout.LabelRole)
                if label_item and label_item.widget():
                    txt = label_item.widget().text()
                    if txt in ("Server IP:", "Remote Path:"):
                        label_item.widget().setVisible(is_sftp)
    def _enforce_one_filter(self, _):
        if self.chk_site.isChecked() and self.chk_model.isChecked():
            sender = self.sender()
            if sender == self.chk_site:
                self.chk_model.setChecked(False)
            else:
                self.chk_site.setChecked(False)

    def _update_ap_device_from_model(self, label: str):
        if label in ("All AP Models", "AP1852/2802/3802/4802"):
            self.ap_device = "cos"
        elif label == "C9105AX/9115AX/9120AX":
            self.ap_device = "cos_bcm"
        else:
            self.ap_device = "cos_qca"

    def _step5_preview(self):

        # ---- WLC Only → filters not allowed ----
        if self.operation_type == "WLC Only":
            self.ap_filter_mode = "NONE"
            self._fill_preview()
            self._goto_step(5)
            return

        # ---- Normal filtering ----
        if self.chk_site.isChecked():
            self.ap_filter_mode = "SITE"
            self.site_tag = self.site_tag_txt.text().strip()
            if not self.site_tag:
                QMessageBox.critical(self, "Missing", "Enter SiteTag Name.")
                return

        elif self.chk_model.isChecked():
            self.ap_filter_mode = "MODEL"
            self.model_group = self.model_dd.currentText()

        else:
            self.ap_filter_mode = "NONE"

        self._fill_preview()
        self._goto_step(5)
    def _sync_iter_state(self):
        """Read iteration widgets into self.iterations_* state variables."""
        try:
            self.iterations_enabled = self.chk_iterations.isChecked()
        except Exception:
            self.iterations_enabled = False
        try:
            v = int(self.iter_count_field.text().strip())
            self.iteration_count = max(1, min(50, v))
        except Exception:
            self.iteration_count = 1
        try:
            v = int(self.iter_interval_field.text().strip())
            self.iteration_interval = max(0, min(18000, v))
        except Exception:
            self.iteration_interval = 300
    def _check_iter_compatibility(self):
        """
        Disable iterations if the current workflow/commands involve
        image download or reload — these are not safe to repeat.
        """
        incompatible_workflows = {"AP Image Download", "TMP Cleanup + reload", "AP Datapath Queue Mon"}
        incompatible_cmd_patterns = (
            "archive download-sw", "sftp://", "scp://", "reload", "%reload%"
        )

        wf = getattr(self, "workflow", "")
        ap_cmds = getattr(self, "ap_cmds", [])

        cmd_has_incompatible = any(
            pat in cmd.lower()
            for cmd in ap_cmds
            for pat in incompatible_cmd_patterns
        )

        should_disable = wf in incompatible_workflows or cmd_has_incompatible

        if hasattr(self, "chk_iterations"):
            if should_disable:
                self.chk_iterations.setChecked(False)
                self.chk_iterations.setEnabled(False)
                self.chk_iterations.setToolTip(
                    "Iterations not available for image download or reload workflows."
                )
            else:
                self.chk_iterations.setEnabled(True)
                self.chk_iterations.setToolTip("")
    def _on_finished(self, summary: dict):
        try:
            # ---------------- BASIC INFO ----------------
            start = summary.get("start")
            end = summary.get("end")
            elapsed = int((end - start).total_seconds()) if start and end else 0

            lines = []
            lines.append(f"Operation Type Selected in Step1: {summary.get('operation', '')}")

            if summary.get("operation") != "AP Only":
                wlc_ip = self.wlc_entries[0]["ip"].text().strip() if getattr(self, "wlc_entries", []) else ""
                if wlc_ip:
                    lines.append(f"WLC IP address: {wlc_ip}")

            if summary.get("operation") in ("WLC & AP", "AP Only"):
                lines.append(f"Total number of APs Processed: {summary.get('ap_total', 0)}")
                lines.append(f"Success APs: {summary.get('ap_success', 0)}")
                lines.append(f"Failed APs: {summary.get('ap_failed', 0)}")

            if summary.get("SiteTagNameFilter"):
                lines.append(
                    f"Selected {summary.get('ApFilteredCnt', 0)} out of {summary.get('TotalApCnt', 0)} total Aps from Site Tag filter '{summary.get('SiteTagNameFilter')}'"
                )

            lines.append(f"Time taken: {elapsed}s")

            if summary.get("operation") == "WLC Only":
                out = summary.get("wlc_output", "")
                lines.append(
                    f"writing outputs to the folder 'data' to the file named {os.path.basename(out) if out else 'eWLC-9800_outputs.txt'}"
                )
            else:
                lines.append(f"Outputs stored in: {summary.get('data_dir', '')}")

            status_file = summary.get("status_summary_file", "")
            if status_file:
                lines.append(f"Status Check Summary File: {status_file}")

            if hasattr(self, "results_summary"):
                try:
                    self.results_summary.setPlainText("\n".join(lines))
                except Exception:
                    pass

            # ---------------- FLAGS ----------------
            wf = summary.get("workflow") or getattr(self, "workflow", None)
            is_wlc_only = summary.get("operation") == "WLC Only"
            # ---------------- CLIENT AUTH LOOP FIXER ----------------
            if "delete_list" in summary:

                delete_list = summary["delete_list"]

                if not delete_list:
                    if hasattr(self, "run_log"):
                        self.run_log.append("[AUTH] No clients to deauthenticate.\n")
                    # ← NO return here — fall through to auto-save below
                else:
                    if hasattr(self, "run_log"):
                        self.run_log.append("\n[AUTH] Auto-deauthenticating stuck clients...\n")

                    engine = PollerEngine(log_cb=lambda m: self.run_log.append(m))
                    engine.deauth_clients(delete_list)

                    if hasattr(self, "run_log"):
                        self.run_log.append("[AUTH] Deauthentication completed\n")

                    # ---- Auto-save run log — runs whether or not delete_list is empty ----
                    if getattr(self, "headless_mode", False):
                        try:
                            folder = str(summary.get("data_dir", DATA_DIR))
                            os.makedirs(folder, exist_ok=True)
                            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                            fn = os.path.join(folder, f"client_auth_loop_logs_{ts}.txt")
                            txt = ""
                            if hasattr(self, "run_log") and self.run_log is not None:
                                try:
                                    txt = self.run_log.toPlainText()
                                except Exception:
                                    txt = ""
                            with open(fn, "w", encoding="utf-8") as _f:
                                _f.write(txt)
                            if hasattr(self, "run_log"):
                                self.run_log.append(f"[AUTH] Run log auto-saved to: {fn}")
                        except Exception as _e:
                            if hasattr(self, "run_log"):
                                self.run_log.append(f"[AUTH] Auto-save failed: {_e}")

                                if hasattr(self, "run_log"):
                                    self.run_log.append("\n[AUTH] Auto-deauthenticating stuck clients...\n")

                                engine = PollerEngine(log_cb=lambda m: self.run_log.append(m))
                                engine.deauth_clients(delete_list)

                                if hasattr(self, "run_log"):
                                    self.run_log.append("[AUTH] Deauthentication completed\n")

                # ---- Auto-save run log for Client Auth Loop workflow ----
                try:
                    folder = str(summary.get("data_dir", DATA_DIR))
                    os.makedirs(folder, exist_ok=True)
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    fn = os.path.join(folder, f"client_auth_loop_logs_{ts}.txt")
                    txt = ""
                    if hasattr(self, "run_log") and self.run_log is not None:
                        try:
                            txt = self.run_log.toPlainText()
                        except Exception:
                            txt = ""
                    with open(fn, "w", encoding="utf-8") as _f:
                        _f.write(txt)
                    if hasattr(self, "run_log"):
                        self.run_log.append(f"[AUTH] Run log auto-saved to: {fn}")
                except Exception as _e:
                    if hasattr(self, "run_log"):
                        self.run_log.append(f"[AUTH] Auto-save failed: {_e}")

            # ---------------- VULNERABLE TABLE ----------------
            default_vuln_headers = ["AP Name", "AP Model", "AP IP", "Active Boot Part", "Recovery", "Partition Note"]
            if hasattr(self, "vuln_table"):
                try:
                    self.vuln_table.setRowCount(0)
                    if wf != "AP Datapath Queue Mon":
                        self.vuln_table.setHorizontalHeaderLabels(default_vuln_headers)
                except Exception:
                    pass

            # ---------------- AP DATAPATH QUEUE MON RESULTS ----------------
            if (not is_wlc_only) and wf == "AP Datapath Queue Mon":
                dp_results = summary.get("datapath_results", [])

                if hasattr(self, "run_log"):
                    try:
                        self.run_log.append("\n===== AP DATAPATH QUEUE MONITOR SUMMARY =====")
                        self.run_log.append(f"Total APs checked : {len(dp_results)}")
                        all_dps = [dp for r in dp_results for dp in r.get("datapaths", [])]
                        healthy = sum(1 for dp in all_dps if dp.get("overall_assessment") == "Healthy")
                        stuck = sum(1 for dp in all_dps if dp.get("overall_assessment") == "Possible Datapath Queue Stuck")
                        no_clients = sum(1 for r in dp_results if not r.get("datapaths"))
                        self.run_log.append(f"Healthy Datapaths      : {healthy}")
                        self.run_log.append(f"Possibly Stuck         : {stuck}")
                        self.run_log.append(f"APs with no clients    : {no_clients}")
                        self.run_log.append("=" * 50)
                        for r in dp_results:
                            self.run_log.append(f"{r.get('ap_name')} ({r.get('ap_ip')}): {r.get('status')}")
                            for dp in r.get("datapaths", []):
                                self.run_log.append(
                                    f"    [{dp.get('datapath_id')}] clients={dp.get('clients')} -> "
                                    f"{dp.get('overall_assessment', '')}"
                                )
                                if dp.get("recommended_recovery"):
                                    self.run_log.append(f"        Recovery: {dp['recommended_recovery']}")
                    except Exception:
                        pass

                if hasattr(self, "results_summary"):
                    try:
                        rs_lines = self.results_summary.toPlainText()
                        rs_lines += (
                            f"\n\nAP Datapath Queue Mon: {len(dp_results)} AP(s) checked | "
                            f"Success: {summary.get('ap_success', 0)} | Failed: {summary.get('ap_failed', 0)}"
                        )
                        self.results_summary.setPlainText(rs_lines)
                    except Exception:
                        pass

                
                if hasattr(self, "vuln_section"):
                    try:
                        self.vuln_section.setVisible(True)
                    except Exception:
                        pass
                if hasattr(self, "ap_section"):
                    try:
                        self.ap_section.setVisible(True)
                    except Exception:
                        pass

                # ---- Populate Susceptible table: only datapaths flagged stuck ----
                if hasattr(self, "vuln_table"):
                    try:
                        self.vuln_table.setHorizontalHeaderLabels(
                            ["AP Name", "AP Model", "AP IP", "Datapath", "Recovery", "Clients"]
                        )
                        for r in dp_results:
                            ap_ip = r.get("ap_ip", "")
                            ap_name = r.get("ap_name", "")
                            for dp in r.get("datapaths", []):
                                if dp.get("overall_assessment") != "Possible Datapath Queue Stuck":
                                    continue

                                real_model = "UNKNOWN"
                                for ap_r in range(self.ap_table.rowCount()):
                                    ip_item = self.ap_table.item(ap_r, 2)
                                    if ip_item and ip_item.text() == ap_ip:
                                        m_item = self.ap_table.item(ap_r, 1)
                                        if m_item and m_item.text():
                                            real_model = m_item.text()
                                        break

                                row_idx = self.vuln_table.rowCount()
                                self.vuln_table.insertRow(row_idx)
                                self.vuln_table.setItem(row_idx, 0, QTableWidgetItem(ap_name))
                                self.vuln_table.setItem(row_idx, 1, QTableWidgetItem(real_model))
                                self.vuln_table.setItem(row_idx, 2, QTableWidgetItem(ap_ip))
                                self.vuln_table.setItem(row_idx, 3, QTableWidgetItem(dp.get("datapath_id", "")))
                                self.vuln_table.setItem(row_idx, 4, QTableWidgetItem(
                                    dp.get("recommended_recovery", "")))
                                self.vuln_table.setItem(row_idx, 5, QTableWidgetItem(
                                    ",".join(dp.get("clients", []))))
                    except Exception as e:
                        if hasattr(self, "run_log"):
                            try:
                                self.run_log.append(f"[DEBUG] Failed populating vuln_table (datapath): {e}")
                            except Exception:
                                pass
                # ---- Save a dedicated report file (separate from per-AP logs) ----
                try:
                    folder = str(summary.get("data_dir", DATA_DIR))
                    os.makedirs(folder, exist_ok=True)
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    fn = os.path.join(folder, f"DatapathQueueMon_Summary_{ts}.txt")
                    with open(fn, "w", encoding="utf-8") as _f:
                        _f.write("=" * 60 + "\n")
                        _f.write("AP DATAPATH QUEUE MONITOR — RUN SUMMARY\n")
                        _f.write("=" * 60 + "\n")
                        _f.write(f"Run time      : {datetime.now().isoformat()}\n")
                        _f.write(f"Client MAC    : {getattr(self, 'datapath_client_mac', '')}\n")
                        _f.write(f"AP list file  : {getattr(self, 'ap_list_file', '')}\n")
                        _f.write(f"Total APs     : {len(dp_results)}\n")
                        _f.write(f"Success       : {summary.get('ap_success', 0)}\n")
                        _f.write(f"Failed        : {summary.get('ap_failed', 0)}\n")
                        _f.write("-" * 60 + "\n\n")
                        for r in dp_results:
                            _f.write(f"AP            : {r.get('ap_name')} ({r.get('ap_ip')})\n")
                            _f.write(f"Status        : {r.get('status')}\n")

                            dps = r.get("datapaths", [])
                            if not dps:
                                _f.write("Datapath ID   : n/a\n")
                                _f.write("Radio ID      : n/a\n")
                                _f.write("Assessment    : n/a\n")
                            else:
                                for dp in dps:
                                    _f.write(f"  Datapath ID : {dp.get('datapath_id', 'n/a')}\n")
                                    _f.write(f"  Radio ID    : {dp.get('radio_id', 'n/a')}\n")
                                    _f.write(f"  Clients     : {','.join(dp.get('clients', []))}\n")
                                    _f.write(f"  Assessment  : {dp.get('overall_assessment', '')}\n")
                                    if dp.get("recommended_recovery"):
                                        _f.write(f"  Recovery    : {dp.get('recommended_recovery')}\n")
                                    _f.write("  " + "-" * 20 + "\n")

                            _f.write("-" * 40 + "\n")
                        _f.write("\nNote: per-AP raw CLI output and per-AP detailed report are in\n")
                        _f.write("DatapathMon_<model>_<apname>.log files in this same folder.\n")
                    if hasattr(self, "run_log"):
                        self.run_log.append(f"\n[DATAPATH] Summary report saved to: {fn}")
                except Exception as _e:
                    if hasattr(self, "run_log"):
                        self.run_log.append(f"[DATAPATH] Failed to save summary report: {_e}")

            if (not is_wlc_only) and (wf == "AP Flash Checker" or summary.get("vulnerable_rows")):
                if hasattr(self, "run_log"):
                    try:
                        self.run_log.append("\n===== PARSER / FLASH CHECK SUMMARY =====")
                    except Exception:
                        pass

                vuln_rows = summary.get("vulnerable_rows", [])
                # Update AP table with correct model from parser

                vuln_count = len(vuln_rows)

                if hasattr(self, "run_log"):
                    self.run_log.append(f"Total Susceptible APs Detected: {vuln_count}")
                    for vr in vuln_rows:
                        self.run_log.append(
                            f"  - {vr.get('ap_name', '')} ({vr.get('ap_ip', '')}) -> "
                            f"{vr.get('recovery', '')}"
                        )
                if hasattr(self, "vuln_table"):
                    try:
                        for vr in vuln_rows:
                            r = self.vuln_table.rowCount()
                            self.vuln_table.insertRow(r)
                            raw_model = vr.get("ap_model", "")
                            # Extract [reason] from model if present, move it to recovery
                            m = re.match(r'^(.*?)(\[.*\])\s*$', raw_model)
                            if m:
                                clean_model = m.group(1).strip()
                                reason = " " + m.group(2)
                            else:
                                clean_model = raw_model
                                reason = ""
                            self.vuln_table.setItem(r, 0, QTableWidgetItem(vr.get("ap_name", "")))
                            # AFTER — look up real model from AP table by IP
                            real_model = clean_model  # fallback to parser model
                            ap_ip = vr.get("ap_ip", "")
                            for ap_r in range(self.ap_table.rowCount()):
                                ip_col = 2 if self.operation_type == "AP Only" else 3
                                ip_item = self.ap_table.item(ap_r, ip_col)
                                if ip_item and ip_item.text() == ap_ip:
                                    model_col = 1 if self.operation_type == "AP Only" else 2
                                    m_item = self.ap_table.item(ap_r, model_col)
                                    if m_item and m_item.text() and m_item.text() != "UNKNOWN":
                                        real_model = m_item.text()
                                    break
                            self.vuln_table.setItem(r, 1, QTableWidgetItem(real_model))
                            self.vuln_table.setItem(r, 2, QTableWidgetItem(vr.get("ap_ip", "")))
                            self.vuln_table.setItem(r, 3, QTableWidgetItem(vr.get("active_boot_part", "")))
                            self.vuln_table.setItem(r, 4, QTableWidgetItem(vr.get("recovery", "") + reason))
                            self.vuln_table.setItem(r, 5, QTableWidgetItem(vr.get("partition_note", "")))
                    except Exception as e:
                        if hasattr(self, "run_log"):
                            try:
                                self.run_log.append(f"[DEBUG] Failed populating vuln_table: {e}")
                            except Exception:
                                pass
                
            # ---------------- AP TABLE (skip for WLC only) ----------------
            if not is_wlc_only and wf != "AP Datapath Queue Mon":
                try:
                    # So AP Only mode never loads old WLC files.
                    need_populate = (
                        hasattr(self, "ap_table")
                        and self.ap_table.rowCount() == 0
                        and self.operation_type != "WLC & AP"  # WLC & AP always uses live signals
                    )

                    if need_populate:
                        cand_files = []
                        try:
                            for fn in os.listdir(CONFD):
                                if fn.lower().startswith("ap_ip_list") and fn.lower().endswith(".txt"):
                                    cand_files.append(os.path.join(CONFD, fn))
                        except Exception:
                            cand_files = []

                        src_file = ""
                        for f in cand_files:
                            if "all" in os.path.basename(f).lower():
                                src_file = f
                                break
                        if not src_file and cand_files:
                            src_file = cand_files[0]

                        if src_file and os.path.exists(src_file):
                            try:
                                if src_file and os.path.exists(src_file):
                                    try:
                                        with open(src_file, "r", encoding="utf-8", errors="ignore") as fh:
                                            # Clear table and populate rows in new order:
                                            # AP Name | AP Model | AP IP | Status
                                            if hasattr(self, "ap_table"):
                                                self.ap_table.setRowCount(0)
                                            rows_added = 0
                                            for i, line in enumerate(fh):
                                                if i >= 200:
                                                    break
                                                s = line.strip()
                                                if not s:
                                                    continue

                                                # split by comma if present (CSV), otherwise by whitespace
                                                parts = [p.strip() for p in (s.split(",") if "," in s else s.split())]

                                                # Basic defaults
                                                ip, model, name = normalize_ap_entry(parts)
                                                # Save mapping so later ap_update can use it
                                                try:
                                                    if name:
                                                        self.ap_name_map[ip] = name
                                                except Exception:
                                                    pass

                                                # Insert into table as: Name | Model | IP | Status
                                                if hasattr(self, "ap_table"):
                                                    r = self.ap_table.rowCount()
                                                    self.ap_table.insertRow(r)
                                                    try:
                                                        self.ap_table.setItem(r, 0, QTableWidgetItem(name))
                                                        self.ap_table.setItem(r, 1, QTableWidgetItem(model))
                                                        self.ap_table.setItem(r, 2, QTableWidgetItem(ip))
                                                        self.ap_table.setItem(r, 3, QTableWidgetItem("Pending"))
                                                    except Exception:
                                                        # fallback: attempt safer sets
                                                        try:
                                                            if not self.ap_table.item(r, 0):
                                                                self.ap_table.setItem(r, 0, QTableWidgetItem(name))
                                                        except Exception:
                                                            pass

                                                rows_added += 1

                                            if hasattr(self, "run_log"):
                                                try:
                                                    self.run_log.append(
                                                        f"[DEBUG] Populated AP Table with {rows_added} rows from {os.path.basename(src_file)}")
                                                except Exception:
                                                    pass
                                    except Exception as e:
                                        if hasattr(self, "run_log"):
                                            try:
                                                self.run_log.append(f"[DEBUG] Exception reading fallback AP list: {e}")
                                            except Exception:
                                                pass

                                            except Exception:
                                                # fallback: attempt safer sets
                                                try:
                                                    if not self.ap_table.item(r, 0): self.ap_table.setItem(r, 0,
                                                                                                           QTableWidgetItem(
                                                                                                               name))
                                                except Exception:
                                                    pass
                                        rows_added += 1

                                if hasattr(self, "run_log"):
                                    try:
                                        self.run_log.append(
                                            f"[DEBUG] Populated AP Table with {rows_added} rows from {os.path.basename(src_file)}")
                                    except Exception:
                                        pass
                            except Exception as e:
                                if hasattr(self, "run_log"):
                                    try:
                                        self.run_log.append(f"[DEBUG] Exception reading fallback AP list: {e}")
                                    except Exception:
                                        pass

                except Exception as e:
                    if hasattr(self, "run_log"):
                        try:
                            self.run_log.append(f"[DEBUG] AP table population error: {e}")
                        except Exception:
                            pass
            # ---------------- FINAL MODEL CORRECTION ----------------
            vuln_rows = summary.get("vulnerable_rows", [])

            #for vr in vuln_rows:
             #   ip = vr.get("ap_ip")
              #  model = vr.get("ap_model")

               # for r in range(self.ap_table.rowCount()):
                #    ip_item = self.ap_table.item(r, 2)
                 #   if ip_item and ip_item.text() == ip:
                  #      if model and model != "UNKNOWN":
                   #         clean_model = re.sub(r'\[.*\]', '', model).strip()
                    #        self.ap_table.setItem(r, 1, QTableWidgetItem(clean_model))
                     #   break


           
        except Exception as e:
            if hasattr(self, "run_log"):
                try:
                    self.run_log.append(f"[ERROR] _on_finished exception: {e}")
                except Exception:
                    pass
            else:
                print("[ERROR] _on_finished exception:", e)

        # finally mark run_in_progress false
        try:
            self.run_count += 1

            ap_table_empty = (
                hasattr(self, "ap_table")
                and self.operation_type != "WLC Only"
                and self.ap_table.rowCount() == 0
            )
            
            if (
                self.run_count >= 2
                and not getattr(self, "headless_mode", False)
                and not ap_table_empty
                and summary.get("ap_total", 0) > 0
            ):
                QMessageBox.information(
                    self,
                    "Restart Recommended",
                    "You have completed multiple runs in the same session.\n\n"
                    "For best results, please save your logs and restart the application before running again.\n\n"
                    "Continuing without restart may cause unexpected behaviour."
                )
            self.run_in_progress = False
            

            # Auto-quit in headless mode after run completes
            if getattr(self, "headless_mode", False) and win.ini.cfg.has_section("CRONJOB"):
                print("[HEADLESS] Run complete. Staying on Step7 for user review.")
                try:
                    self._save_run_log_headless()
                except Exception:
                    pass
                # Stay on Step7 — user decides via Delete CronJob button whether to keep it
            try:
                if hasattr(self, "sidebar"):
                    self.sidebar.setEnabled(True)
            except Exception:
                pass
                        
        except Exception:
            pass
    def _save_auth_log_file(self):
        try:
            if not hasattr(self, "run_log"):
                return

            txt = self.run_log.toPlainText().strip()
            if not txt:
                self.run_log.append("[AUTH] No content to save")
                return

            folder = DATA_DIR
            os.makedirs(folder, exist_ok=True)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fn = os.path.join(folder, f"client_auth_loop_logs_{ts}.txt")

            with open(fn, "w", encoding="utf-8") as f:
                f.write(txt)

            self.run_log.append(f"[AUTH] Run log saved to: {fn}")

        except Exception as e:
            self.run_log.append(f"[AUTH] Save failed: {e}")
    def _on_ap_update(self, *args):
        

        # ✅ SAFE UNPACKING (handles both 4 and 6 args)

        row = None
        ip = model = status = name = wlc = ""

        if len(args) == 6:
            row, ip, model, status, name, wlc = args

        elif len(args) == 5:
            row, ip, model, status, name = args

        elif len(args) == 4:
            ip, model, status, name = args

        else:
            print("[ERROR] Unexpected args in _on_ap_update:", args)
            return
        # 🔥 FIX 1: NEVER trust engine row index (prevents overwrite bug)
        row = self.ap_table.rowCount()
        self.ap_table.insertRow(row)
        self.ap_table.resizeRowToContents(row)
        # ---------------- AP ONLY MODE ----------------
        if self.operation_type == "AP Only":

            # fallback name
            if not name:
                name = self.ap_name_map.get(ip, f"AP_{ip.replace('.', '_')}")

            self.ap_table.setItem(row, 0, QTableWidgetItem(name))
            self.ap_table.setItem(row, 1, QTableWidgetItem(model or "UNKNOWN"))
            self.ap_table.setItem(row, 2, QTableWidgetItem(ip))
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            self.ap_table.setItem(row, 3, status_item)

        # ---------------- WLC & AP MODE ----------------
        else:

            self.ap_table.setItem(row, 0, QTableWidgetItem(wlc or "-"))
            self.ap_table.setItem(row, 1, QTableWidgetItem(name or "-"))
            self.ap_table.setItem(row, 2, QTableWidgetItem(model or "UNKNOWN"))
            self.ap_table.setItem(row, 3, QTableWidgetItem(ip))
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            self.ap_table.setItem(row, 4, status_item)

        # 🔥 FIX 2: Safe status coloring
        status_col = self.ap_table.columnCount() - 1
        item = self.ap_table.item(row, status_col)

        if item:
            status_lower = status.lower()

            if "success" in status_lower:
                item.setForeground(Qt.green)
                item.setBackground(Qt.transparent)

            elif "fail" in status_lower or "error" in status_lower:
                item.setForeground(Qt.red)

                # 🔥 restore old highlighted look
                item.setBackground(QColor(255, 230, 230))  # light red background
                item.setToolTip(status)  # full message on hover

        # 🔥 FIX 3: Always scroll (better UX)
        self.ap_table.scrollToBottom()
    def _fill_preview(self):
        """
        Build Step-6 Preview safely.
        Never throws exceptions.
        """

        def get_txt(widget):
            try:
                return widget.text().strip()
            except Exception:
                return ""

        def get_dd(widget):
            try:
                return widget.currentText().strip()
            except Exception:
                return ""

        lines = []
        sec = 1

        def section(title):
            nonlocal sec
            lines.append(f"{sec}) {title}")
            sec += 1

        # ---------------- Operation ----------------
        section("Operation Type")

        op = getattr(self, "operation_type", "")
        lines.append(f"   - {op}")

        if op != "AP Only" and getattr(self, "wlc_entries", []):
            for entry in self.wlc_entries:
                ip = entry["ip"].text().strip()
                if ip:
                    lines.append(f"   - WLC IP: {ip}")

        # ---------------- Credentials ----------------
        lines.append("")
        section("Credentials")

        if op != "AP Only" and getattr(self, "wlc_entries", []):
            lines.append(f"   - WLC Username: {self.wlc_entries[0]['user'].text().strip()}")

        if op != "WLC Only" and hasattr(self, "ap_user"):
            lines.append(f"   - AP Username: {get_txt(self.ap_user)}")

        # ---------------- Workflow ----------------
        # ---------------- Workflow ----------------
        lines.append("")
        section("Workflow")

        op = getattr(self, "operation_type", "")

        # For WLC Only operation → fixed workflow
        if op == "WLC Only":
            wf = "Custom WLC CLI Commands"
        else:
            wf = getattr(self, "workflow", "")
            if hasattr(self, "workflow_dd"):
                wf = get_dd(self.workflow_dd)

        lines.append(f"   - {wf}")
        if self.workflow == "Client Stuck In Auth Loop":

            lines.append("Workflow: Client Stuck In Auth Loop")
            lines.append("")
            lines.append("Operation Summary:")
            lines.append("- Detect clients stuck in 'Authenticating' state by:")
            lines.append("    1. Compare client list over 2 intervals (15 mins apart)")
            lines.append("    2. Identify persistent/common clients stuck in Authenticating state")
            lines.append("- Collect debugs including Archive & WNCD Core")
            lines.append("- Dump 'show wireless client mac detail' for clients to be deleted/recovered to file")
            lines.append("- Recover the client by deauthenticating affected clients automatically")
            lines.append("")
            lines.append("Commands Used:")
            lines.append("1. show wireless client summary | include Authenticating")
            lines.append("2. Sleep for 15 mins")
            lines.append("3. show wireless client summary | include Authenticating")
            lines.append("4. Take the common clients in above 2 samples to prepare the delete list")

            lines.append("5. request platform software trace archive last 1 hour target bootflash:<file>")
            lines.append("6. request platform software process core wncd 0 chassis active r0")

            lines.append("7. show wireless client mac-address <mac> detail  (for each client in delete list)")
            lines.append("8. wireless client mac-address <mac> deauthenticate")
        # ---------------- CLI COMMANDS  (MERGED) ----------------
        wlc_cmds = []
        ap_cmds = []

        if op != "AP Only":
            wlc_cmds = getattr(self, "wlc_cmds", [])
            if not wlc_cmds and hasattr(self, "wlc_cmd_box"):
                wlc_cmds = [c.strip() for c in self.wlc_cmd_box.toPlainText().splitlines() if c.strip()]

        if op != "WLC Only":
            ap_cmds = getattr(self, "ap_cmds", [])
            if not ap_cmds and hasattr(self, "ap_cmd_box"):
                ap_cmds = [c.strip() for c in self.ap_cmd_box.toPlainText().splitlines() if c.strip()]

        if wlc_cmds or ap_cmds:
            lines.append("")
            section("CLI Commands")

            if wlc_cmds:
                lines.append("   WLC:")
                for c in wlc_cmds:
                    lines.append(f"      • {c}")

            if ap_cmds:
                lines.append("   AP:")
                for c in ap_cmds:
                    lines.append(f"      • {c}")


        # ---------------- AP Filters ----------------
        # ---------------- Upload Settings ----------------
        if getattr(self, "workflow", "") == "Upload Files from AP":
            lines.append("")
            section("Upload Settings")
            ft  = self.upload_file_type_dd.currentText()     if hasattr(self, "upload_file_type_dd")     else getattr(self, "upload_file_type", "")
            pr  = self.upload_proto_dd.currentText()         if hasattr(self, "upload_proto_dd")         else getattr(self, "upload_proto", "TFTP")
            srv = self.upload_server_ip_field.text().strip() if hasattr(self, "upload_server_ip_field") else getattr(self, "upload_server_ip_val", "")
            lines.append(f"   - File Type : {ft}")
            lines.append(f"   - Protocol  : {pr}")
            lines.append(f"   - Server IP : {srv}")

        # ---------------- AP Filters ----------------
        lines.append("")
        section("AP Filters")

        if op == "WLC Only":
            lines.append("   - Not Applicable")
        else:
            mode = getattr(self, "ap_filter_mode", "NONE")

            if mode == "SITE":
                site = getattr(self, "site_tag", "")
                if hasattr(self, "site_tag_txt"):
                    site = get_txt(self.site_tag_txt) or site
                lines.append(f"   - Site Tag: {site}")

            elif mode == "MODEL":
                model = getattr(self, "model_group", "")
                if hasattr(self, "model_dd"):
                    model = get_dd(self.model_dd)
                lines.append(f"   - Model Group: {model}")

            else:
                lines.append("   - No Filters Applied")
        if self.workflow == "TMP Cleanup + reload":
            preview = []
            preview.append("Workflow: AP Cleanup + Reload")
            preview.append("")

            preview.append("  1. Connects to each AP via SSH")
            preview.append("  2. Executes 'reload' command on each AP")
            preview.append("  3. Automatically confirms reload (if prompted)")
            preview.append("")

            preview.append("Commands executed per AP:")
            preview.append("  • reload")
            preview.append("")
            preview.append("⚠ WARNING:")
            preview.append("  This operation will reboot ALL selected APs.")
            preview.append("  Ensure maintenance window before proceeding.")
            op = getattr(self, "operation_type", "")
            if op == "AP Only":
                preview.append(f"AP List File: {getattr(self, 'ap_list_file', '')}")
            else:
                wlc_ip = self.wlc_entries[0]["ip"].text().strip() if getattr(self, "wlc_entries", []) else ""
                preview.append(f"WLC IP: {wlc_ip}")
            if hasattr(self, "preview_text"):
                self.preview_text.setPlainText("\n".join(preview))
            return

        
        # ---------------- Render Preview ----------------
        if hasattr(self, "preview_text"):
            try:
                self.preview_text.setPlainText("\n".join(lines))
            except Exception:
                pass
        # ── ITERATION CONFIG ─────────────────────────────────────
        lines.append("")
        lines.append(f"{sec}) Iteration Config")
        lines.append("   " + "-" * 30)
        enabled = getattr(self, "iterations_enabled", False)
        lines.append(f"   - Enabled  "       if enabled else "   -Disabled")
        if enabled:
            lines.append(f"   - Count          : {getattr(self, 'iteration_count', 1)}")
            lines.append(f"   - Interval (sec) : {getattr(self, 'iteration_interval', 300)}")

        if hasattr(self, "preview_text"):
            try:
                self.preview_text.setPlainText("\n".join(lines))
            except Exception:
                pass        
    def _run_parser(self):
        """
        Run parser/search over the most recent 'data' leaf folder.
        Supports regex (preferred) or plain substring if regex compilation fails.
        Only searches files appropriate for the selected parser mode:
          - "WLC files" => only files ending with "_outputs.txt"
          - "AP files"  => everything except files ending with "_outputs.txt"
        Results are written to self.parser_out (QTextEdit).
        Defensive: checks attributes before using them.
        """
        try:
            # get pattern text
            if not hasattr(self, "parser_pattern") or self.parser_pattern is None:
                QMessageBox.critical(self, "Missing", "Parser input control not found.")
                return
            pat_text = self.parser_pattern.text().strip()
            if not pat_text:
                QMessageBox.critical(self, "Missing", "Enter a pattern to search.")
                return

            # determine latest data folder
            data_root = DATA_DIR
            if not os.path.exists(data_root):
                # nothing to search
                out = "No 'data' folder found."
                if hasattr(self, "parser_out"):
                    self.parser_out.setPlainText(out)
                else:
                    print(out)
                return

            leafs = []
            for top, dirs, files in os.walk(data_root):
                if files:
                    leafs.append(top)
            latest = max(leafs) if leafs else data_root

            # choose mode
            use_wlc = False
            if hasattr(self, "parser_mode") and callable(getattr(self.parser_mode, "currentText", None)):
                use_wlc = (self.parser_mode.currentText() == "WLC files")
            else:
                # fallback if parser_mode is stored somewhere else
                use_wlc = getattr(self, "parser_mode_value", "WLC files") == "WLC files"

            out_lines = [f"Searching in: {latest}", f"Mode: {'WLC files' if use_wlc else 'AP files'}",
                         f"Pattern: {pat_text}", ""]

            # try compile regex; if it fails, fall back to substring search but report the regex error
            rx = None
            try:
                rx = re.compile(pat_text)
            except Exception as e:
                out_lines.append(
                    f"[Note] Pattern is not a valid regex — falling back to substring search. Regex error: {e}")
                out_lines.append("")

            found_any = False
            try:
                files = sorted(os.listdir(latest))
            except Exception as e:
                out_lines.append(f"[ERROR] Unable to list files in {latest}: {e}")
                if hasattr(self, "parser_out"):
                    self.parser_out.setPlainText("\n".join(out_lines))
                return

            for fn in files:
                # filter files depending on mode
                if use_wlc and not fn.endswith("_outputs.txt"):
                    continue
                if (not use_wlc) and fn.endswith("_outputs.txt"):
                    continue

                fp = os.path.join(latest, fn)
                try:
                    with open(fp, "r", encoding="utf-8", errors="ignore") as f:
                        txt = f.read()
                    match = False
                    if rx:
                        if rx.search(txt):
                            match = True
                    else:
                        if pat_text in txt:
                            match = True
                    if match:
                        found_any = True
                        out_lines.append(f"[MATCH] {fn}")
                except Exception as e:
                    out_lines.append(f"[ERROR] {fn}: {e}")

            if not found_any:
                out_lines.append("No matches found.")

            # write to parser_out if present
            if hasattr(self, "parser_out") and self.parser_out is not None:
                self.parser_out.setPlainText("\n".join(out_lines))
            else:
                print("\n".join(out_lines))

        except Exception as e:
            # last-resort error handling
            try:
                QMessageBox.critical(self, "Parser Error", f"Unexpected error while parsing: {e}")
            except Exception:
                print("Unexpected error in _run_parser:", e)

    def _refresh_visibility(self):

        # ---------------- STEP 1 ----------------
        if hasattr(self, "ap_upload_row"):
            self.ap_upload_row.setVisible(self.operation_type == "AP Only")
        elif hasattr(self, "ap_upload_box"):
            self.ap_upload_box.setVisible(self.operation_type == "AP Only")

        if hasattr(self, "btn_step1_next"):
            self.btn_step1_next.setEnabled(True)

            if self.operation_type == "AP Only":
                self.btn_step1_next.setToolTip(
                    "Please upload an AP list file before proceeding."
                )
            else:
                self.btn_step1_next.setToolTip("")

        # ---------------- STEP 2 (FIXED) ----------------
        show_wlc = self.operation_type in ("WLC Only", "WLC & AP")
        show_ap = self.operation_type in ("AP Only", "WLC & AP")

        if hasattr(self, "wlc_block"):
            self.wlc_block.setVisible(show_wlc)

        if hasattr(self, "ap_block"):
            self.ap_block.setVisible(show_ap)

        # ---------------- STEP 4 (FIXED) ----------------
        if hasattr(self, "ap_cmd_section"):
            self.ap_cmd_section.setVisible(self.operation_type != "WLC Only")

        # ---------------- COMMAND BOX VISIBILITY ----------------
        ap_mode_text = (
            self.ap_mode_dd.currentText()
            if hasattr(self, "ap_mode_dd")
            else getattr(self, "ap_mode", "AP Custom Cmd List")
        )

        show_wlc_cmd = (
                self.operation_type in ("WLC Only", "WLC & AP")
                and getattr(self, "workflow", "") not in ("AP Flash Checker", "Upload Files from AP")
        )
        if hasattr(self, "wlc_cmd_section"):
            self.wlc_cmd_section.setVisible(show_wlc_cmd)
        elif hasattr(self, "wlc_cmd_box"):
            self.wlc_cmd_box.setVisible(show_wlc_cmd)
        show_ap_cmd = (
                              self.operation_type in ("AP Only", "WLC & AP")
                      ) or getattr(self, "workflow", "") == "AP Flash Checker"
        if hasattr(self, "ap_cmd_box"):
            self.ap_cmd_box.setVisible(show_ap_cmd)
        # ---------------- FTP ----------------
        ftp_visible = (
                self.operation_type in ("AP Only", "WLC & AP")
                and ap_mode_text == "AP Image Download"
        )
        if hasattr(self, "ftp_group"):
            self.ftp_group.setVisible(ftp_visible)

        # ---------------- FILTERS ----------------
        filters_allowed = self.operation_type in ("WLC & AP", "AP Only")
        # Disable model filter for AP Flash Checker in WLC & AP mode
        if (
                self.operation_type == "WLC & AP"
                and getattr(self, "workflow", "") == "AP Flash Checker"
        ):
            if hasattr(self, "chk_model"):
                self.chk_model.setChecked(False)
                self.chk_model.setVisible(False)

            if hasattr(self, "model_dd"):
                self.model_dd.setVisible(False)
        else:
            # Restore visibility for other workflows
            if hasattr(self, "chk_model"):
                self.chk_model.setVisible(True)

            if hasattr(self, "model_dd"):
                self.model_dd.setVisible(True)
        if hasattr(self, "chk_site") and hasattr(self, "chk_model"):
            if not filters_allowed:
                self.chk_site.setChecked(False)
                self.chk_model.setChecked(False)

        # Force Qt layout recalculation
        if self.centralWidget():
            self.centralWidget().updateGeometry()

    def _post_init_layout_fix(self):
        # first apply visibility rules
        self._refresh_visibility()

        # go to step0 to ensure consistent base layout
        self.stack.setCurrentIndex(0)

        # allow Qt to finish geometry

    def _ui_progress_update(self, pct):
        if hasattr(self, "progress"):
            self.progress.setValue(pct)

    def _on_wlc_progress(self, done: int, total: int):
        """Bulk WLC mode only (>3 WLCs) — updates the '<done> / <total> WLCs
        completed' note under the progress bar. Never emitted for the manual
        (<=3 WLC) path, so this never runs outside bulk-mode runs."""
        if hasattr(self, "wlc_progress_label"):
            self.wlc_progress_label.setText(f"{done} / {total} WLCs completed")
            self.wlc_progress_label.setVisible(True)

    def changeEvent(self, event):
        from PySide6.QtCore import QEvent

        if event.type() == QEvent.WindowStateChange:
            if not self.isMinimized():
                if hasattr(self, "ap_table"):
                    self.ap_table.resizeRowsToContents()

        super().changeEvent(event)
    def _add_wlc_entry(self):
        if len(self.wlc_entries) >= 3:
            return  # max 3 WLCs

        idx = len(self.wlc_entries)

        entry_widget = QWidget()
        entry_widget.setStyleSheet("background:#f9fafb; border:1px solid #e5e7eb; border-radius:6px;")
        entry_layout = QVBoxLayout(entry_widget)
        entry_layout.setContentsMargins(10, 8, 10, 8)
        entry_layout.setSpacing(4)

        # Header row: label + remove button
        hdr = QHBoxLayout()
        lbl = QLabel(f"WLC {idx + 1}")
        lbl.setStyleSheet("font-weight:600;")
        hdr.addWidget(lbl)
        hdr.addStretch()

        remove_btn = QPushButton("Remove WLC")
        remove_btn.setFixedHeight(28)
        remove_btn.setVisible(True)
        remove_btn.setStyleSheet("min-width:60px; font-size:11px;")
        remove_btn.clicked.connect(lambda _, ew=entry_widget: self._remove_wlc_entry(ew))
        # Only show remove if not WLC1
        remove_btn.setVisible(idx > 0)
        hdr.addWidget(remove_btn)
        entry_layout.addLayout(hdr)

        form = QFormLayout()
        form.setSpacing(4)

        ip_field = QLineEdit()
        ip_field.setFixedHeight(28)
        ip_field.setPlaceholderText("WLC IP  (Required)" if idx > 0 else "WLC IP")

        user_field = QLineEdit()
        user_field.setFixedHeight(28)
        user_field.setPlaceholderText("Username  (Required)" if idx > 0 else "Username")

        pasw_field = QLineEdit()
        pasw_field.setFixedHeight(28)
        pasw_field.setEchoMode(QLineEdit.Password)
        pasw_field.setPlaceholderText("Password  (Required)" if idx > 0 else "Password")

        # Red border on empty mandatory fields for WLC2/WLC3
        if idx > 0:
            EMPTY_STYLE  = "border: 1px solid #dc2626; border-radius:4px;"
            FILLED_STYLE = "border: 1px solid #e6e8eb; border-radius:4px;"

            def _update_style(text, field=None):
                field.setStyleSheet(EMPTY_STYLE if not text.strip() else FILLED_STYLE)

            ip_field.textChanged.connect(lambda t, f=ip_field: _update_style(t, f))
            user_field.textChanged.connect(lambda t, f=user_field: _update_style(t, f))
            pasw_field.textChanged.connect(lambda t, f=pasw_field: _update_style(t, f))

            # Set initial red border since fields start empty
            ip_field.setStyleSheet(EMPTY_STYLE)
            user_field.setStyleSheet(EMPTY_STYLE)
            pasw_field.setStyleSheet(EMPTY_STYLE)

        # Pre-fill from ini
        if self.ini:
            section = "WLC" if idx == 0 else f"WLC{idx+1}"
            ip_field.setText(self.ini.get(section, "wlc_ip"))
            user_field.setText(self.ini.get(section, "wlc_user"))
            pasw_field.setText(self.ini.get(section, "wlc_pasw"))

        form.addRow("IP:", ip_field)
        ip_row_label = form.labelForField(ip_field)
        form.addRow("User:", user_field)
        form.addRow("Password:", pasw_field)
        entry_layout.addLayout(form)

        self.wlc_entries_layout.addWidget(entry_widget)
        self.wlc_entries.append({
            "widget": entry_widget,
            "ip": ip_field,
            "ip_label": ip_row_label,
            "user": user_field,
            "pasw": pasw_field,
            "remove_btn": remove_btn,
        })

        # Disable Add WLC button when at max
        if hasattr(self, "btn_add_wlc"):
            self.btn_add_wlc.setEnabled(len(self.wlc_entries) < 3)
        if hasattr(self, "per_wlc_cmd_section"):
            self._build_per_wlc_cmd_boxes()
    def _remove_wlc_entry(self, entry_widget):
            if len(self.wlc_entries) <= 1:
                return
            self.wlc_entries = [e for e in self.wlc_entries if e["widget"] is not entry_widget]
            entry_widget.setParent(None)
            entry_widget.deleteLater()

            # Renumber labels and fix remove button visibility
            # Renumber labels and fix remove button visibility
            for i, e in enumerate(self.wlc_entries):
                lbl = e["widget"].findChild(QLabel)
                if lbl:
                    lbl.setText(f"WLC {i + 1}")
                btn = e.get("remove_btn")
                if btn:
                    btn.setText("Remove")   # ensure label is always present
                    btn.setVisible(i > 0)   # WLC1 never shows remove

            # Re-enable Add WLC button since we're below max
            if hasattr(self, "btn_add_wlc"):
                self.btn_add_wlc.setEnabled(len(self.wlc_entries) < 3)
    def _validate_all_wlcs(self):
        missing = []

        for idx, entry in enumerate(self.wlc_entries):
            ip = entry["ip"].text().strip()
            user = entry["user"].text().strip()
            pwd = entry["pasw"].text().strip()   # ✅ FIXED

            if not ip or not user or not pwd:
                missing.append(f"WLC {idx + 1}")

        return missing
    def _remove_last_wlc_entry(self):
        if len(self.wlc_entries) > 1:
            entry = self.wlc_entries.pop()
            entry["widget"].setParent(None)
            entry["widget"].deleteLater()

            # Renumber labels and fix remove button visibility
            for i, e in enumerate(self.wlc_entries):
                lbl = e["widget"].findChild(QLabel)
                if lbl:
                    lbl.setText(f"WLC {i + 1}")
                btn = e.get("remove_btn")
                if btn:
                    btn.setVisible(i > 0)

            # Re-enable Add WLC button since we're below max
            if hasattr(self, "btn_add_wlc"):
                self.btn_add_wlc.setEnabled(len(self.wlc_entries) < 3)
            if hasattr(self, "per_wlc_cmd_section"):
                self._build_per_wlc_cmd_boxes()

    def _on_multi_wlc_toggle(self, checked: bool):
        """UI-only toggle for Bulk Upload mode (more than 3 WLCs). Disables the
        manual Add/Remove WLC controls and swaps in the Excel-based WLC list
        upload option instead. File parsing/validation is not wired up yet."""
        # Switching back to Manual with a file already selected would silently
        # discard it — confirm first, and stay in Bulk mode if the user backs out.
        if not checked and getattr(self, "wlc_bulk_path", None) and self.wlc_bulk_path.text().strip():
            reply = QMessageBox.question(
                self, "Discard Uploaded WLC List?",
                "Switching back to manual entry will discard the uploaded WLC list. Continue?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                self.wlc_mode_bulk_radio.blockSignals(True)
                self.wlc_mode_bulk_radio.setChecked(True)
                self.wlc_mode_bulk_radio.blockSignals(False)
                return
            self.wlc_bulk_path.clear()
            if hasattr(self, "wlc_bulk_stats"):
                self.wlc_bulk_stats.setText("")
                self.wlc_bulk_stats.setVisible(False)

        self.btn_add_wlc.setEnabled(not checked)
        self.remove_btn_wlc.setEnabled(not checked)

        if hasattr(self, "wlc_bulk_upload_row"):
            self.wlc_bulk_upload_row.setVisible(checked)
        if hasattr(self, "wlc_multi_note"):
            self.wlc_multi_note.setVisible(checked)

        # Only WLC 1 remains visible — it now holds the shared credentials
        for i, entry in enumerate(self.wlc_entries):
            if i > 0:
                entry["widget"].setVisible(not checked)

        # WLC 1: relabel to make its new purpose explicit, and remove its IP
        # field — the WLC list (with IPs) now comes from the uploaded file.
        if self.wlc_entries:
            wlc1 = self.wlc_entries[0]
            lbl = wlc1["widget"].findChild(QLabel)
            if lbl:
                lbl.setText("Shared WLC Credentials" if checked else "WLC 1")
            wlc1["ip"].setVisible(not checked)
            if wlc1.get("ip_label"):
                wlc1["ip_label"].setVisible(not checked)

    def _browse_wlc_list_excel(self):
        """File picker + parser for the bulk WLC list (Excel, or Text/CSV).

        Supports two formats, auto-detected, columns in any order:
          1) IP only     — a single column of WLC IP addresses
          2) IP + Name    — two columns: WLC IP and WLC Name

        Column identity is resolved from the header row when one is present
        (a cell containing 'ip' / 'name'); otherwise it falls back to
        inspecting the first data row and treating whichever cell parses as
        a valid IPv4 address as the IP column, and any other non-empty cell
        as the name column.

        Excel files (.xlsx/.xls) are read via openpyxl. Text/CSV files
        (.txt/.csv) are read as one entry per line, comma-or-whitespace
        separated (same convention as the AP list upload).
        """
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Upload WLC List File (Format: WLC IP, WLC Name)",
            "",
            "Excel/Text/CSV (*.xlsx *.xls *.txt *.csv);;Excel Files (*.xlsx *.xls);;"
            "Text/CSV (*.txt *.csv);;All Files (*.*)"
        )
        if not path:
            return

        ext = os.path.splitext(path)[1].lower()

        if ext in (".txt", ".csv"):
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    rows = [
                        [p.strip() for p in (line.split(",") if "," in line else line.split())]
                        for line in f
                        if line.strip()
                    ]
            except Exception as e:
                QMessageBox.critical(self, "Invalid File", f"Could not read file:\n{e}")
                return
        else:
            if load_workbook is None:
                QMessageBox.critical(self, "Excel Support Missing",
                                      "openpyxl is not available — cannot read Excel files.")
                return
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = [
                    [("" if c is None else str(c).strip()) for c in row]
                    for row in ws.iter_rows(values_only=True)
                ]
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Invalid File", f"Could not read Excel file:\n{e}")
                return

        rows = [r for r in rows if any(cell for cell in r)]
        if not rows:
            QMessageBox.critical(self, "Invalid File", "No data found in the file.")
            return

        def _is_ip(v: str) -> bool:
            try:
                socket.inet_pton(socket.AF_INET, v)
                return True
            except Exception:
                return False

        # ---- Detect header row & column positions ('ip' / 'name', any order) ----
        ip_col, name_col = None, None
        header_lower = [c.lower() for c in rows[0]]

        if any("ip" in c for c in header_lower):
            for i, c in enumerate(header_lower):
                if "ip" in c and ip_col is None:
                    ip_col = i
                elif "name" in c and name_col is None:
                    name_col = i
            data_rows = rows[1:]
        else:
            # No header — infer columns from the first data row's content
            data_rows = rows
            for i, c in enumerate(rows[0]):
                if _is_ip(c) and ip_col is None:
                    ip_col = i
                elif c and ip_col != i and name_col is None:
                    name_col = i
            if ip_col is None:
                ip_col = 0  # fall back to first column as IP

        total_cnt = valid_cnt = invalid_cnt = duplicate_cnt = 0
        seen_ips = set()
        parsed = []

        for r in data_rows:
            if not any(r):
                continue
            total_cnt += 1
            ip = r[ip_col].strip() if ip_col < len(r) else ""
            name = r[name_col].strip() if (name_col is not None and name_col < len(r)) else ""

            if not _is_ip(ip):
                invalid_cnt += 1
                continue
            if ip in seen_ips:
                duplicate_cnt += 1
                continue

            seen_ips.add(ip)
            valid_cnt += 1
            parsed.append({"ip": ip, "name": name})

        if valid_cnt == 0:
            QMessageBox.critical(self, "Invalid File", "No valid WLC IP addresses found.")
            return

        self.wlc_bulk_path.setText(path)
        # Parsed WLC list — consumed by the execution backend (to be wired up separately)
        self.wlc_bulk_list = parsed

        stats_html = f"""
        <span style='color:#2563eb; font-weight:600;'>Total:</span> {total_cnt} |
        <span style='color:#16a34a; font-weight:600;'>Valid:</span> {valid_cnt} |
        <span style='color:#dc2626; font-weight:600;'>Invalid:</span> {invalid_cnt} |
        <span style='color:#f59e0b; font-weight:600;'>Duplicates:</span> {duplicate_cnt}
        """
        if hasattr(self, "wlc_bulk_stats"):
            self.wlc_bulk_stats.setText(stats_html)
            self.wlc_bulk_stats.setVisible(True)

        QMessageBox.information(
            self, "WLC List Loaded",
            f"""File loaded successfully.

    Total rows: {total_cnt}
    Valid WLCs: {valid_cnt}
    Invalid entries: {invalid_cnt}
    Duplicate IPs: {duplicate_cnt}
    """
        )
    def _update_workflow_dropdown(self):

        if not hasattr(self, "workflow_dd"):
            return

        self.workflow_dd.blockSignals(True)
        self.workflow_dd.clear()

        # ✅ WLC ONLY
        if self.operation_type == "WLC Only":
            self.workflow_dd.addItems([
                "Custom CLI Commands",
                "Client Stuck In Auth Loop"
            ])

        # ✅ WLC & AP  ❗ REMOVE AUTH LOOP HERE
        elif self.operation_type == "WLC & AP":
            self.workflow_dd.addItems([
                "AP Flash Checker",
                "Custom CLI Commands",
                "TMP Cleanup + reload",
                "AP Datapath Queue Mon",     
            ])

        # ✅ AP ONLY
        elif self.operation_type == "AP Only":
            self.workflow_dd.addItems([
                "AP Flash Checker",
                "Custom CLI Commands",
                "Upload Files from AP",
                "TMP Cleanup + reload",
                "AP Datapath Queue Mon",
                
            ])

        self.workflow_dd.setCurrentIndex(0)
        self.workflow_dd.blockSignals(False)
        self._on_workflow_change(self.workflow_dd.currentText())
    def _init_workflow_ui(self):
        self._update_workflow_dropdown()

        # 🔥 CRITICAL: trigger visibility AFTER dropdown is populated
        current = self.workflow_dd.currentText()
        self._on_workflow_change(current)
def resource_path(relpath: str) -> str:
    """
    Return a filesystem path to `relpath` that works when running normally
    and when bundled by PyInstaller.
    Example: resource_path("assets/cisco_logo.ico")
    """
    if getattr(sys, "frozen", False):
        # Running in a PyInstaller bundle
        base = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
    else:
        base = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base, relpath)




# REPLACE WITH:
def main():
    app = QApplication(sys.argv)
    apply_global_style(app)
    app.setWindowIcon(QIcon(resource_path("assets/ciscologo.ico")))
    win = MainWindow()
    app.aboutToQuit.connect(win._stop_worker)
    win.show()

    # ── Auto-run saved CronJob if present in config.ini ──
    if win.ini and win.ini.cfg.has_section("CRONJOB"):
        win.headless_mode = True
        print("[CRONJOB] CronJob found in config.ini — auto-starting...")
        QTimer.singleShot(300, win._load_and_run_cronjob)

    sys.exit(app.exec())




if __name__ == "__main__":
    main()
